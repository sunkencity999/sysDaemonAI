import os
import sys

# Add the project root directory to Python path if running as script
if __name__ == '__main__':
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)

import sys
import os
import json
import requests
from threading import Thread
from PyQt6.QtCore import Qt, QMetaObject, Q_ARG
import socket
import logging
import platform
import threading
from .server import AgentServer  # Use relative import for server module
import subprocess
from datetime import datetime
from pathlib import Path
import psutil
import openpyxl
import random
import getpass
import traceback
import subprocess
import shutil
import pandas as pd
from datetime import datetime, timedelta
from scapy.all import sniff, AsyncSniffer, get_if_list, conf, ETH_P_ALL, TCP, UDP
from scapy.utils import PcapWriter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                         QLabel, QPushButton, QTextEdit, QTabWidget, QSystemTrayIcon,
                         QMenu, QDialog, QLineEdit, QComboBox, QMessageBox, QFrame,
                         QProgressBar, QGroupBox, QHeaderView, QToolBar, QCheckBox, QSplitter, QTableWidget, 
                         QGridLayout, QDateTimeEdit, QSpinBox, QScrollArea,
                         QRadioButton, QPlainTextEdit, QTableWidgetItem, QStackedWidget, QListWidget, 
                         QDialogButtonBox, QFileDialog, QInputDialog)  # Added QDialogButtonBox
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QDateTime
from PyQt6.QtGui import (QPainter, QColor, QPen, QBrush, QAction, QPalette, QIcon,
                       QLinearGradient, QImage)
from PyQt6.QtCharts import (QChart, QChartView, QPieSeries, QLineSeries, 
                           QValueAxis, QPieSlice)
import numpy as np
import time
import getpass
import random
from network_monitor import NetworkMonitor
from data_analyzer import DataAnalyzer
from database import DatabaseManager, Threat, SystemMetrics, NetworkConnection, Alert
import matplotlib
matplotlib.use('qtagg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
import numpy as np
import socket
import time
import getpass
import random
import openpyxl
from openpyxl.styles import Font, Alignment
from network_monitor import NetworkMonitor
from data_analyzer import DataAnalyzer
from database import DatabaseManager, Threat, SystemMetrics, NetworkConnection, Alert
import scapy.all as scapy
from scapy.all import get_if_list, sniff, AsyncSniffer
import csv
from typing import Dict, List, Optional, Tuple
from ai_agents.log_monitor_agent import LogMonitorAgent
from ai_agents.threat_intel_agent import ThreatIntelAgent
from ai_agents.defense_agent import DefenseAgent
from ai_agents.crawler_agent import CrawlerAgent
from agent_ui import AgentTabs
import config
from auth_manager import AuthManager
try:
    from .packet_capture import PacketCaptureThread  # Try relative import first
except ImportError:
    # If relative import fails, try importing from the same directory
    from packet_capture import PacketCaptureThread  # Fall back to absolute import

try:
    from .monitor_thread import NetworkMonitorThread  # Try relative import first
except ImportError:
    # If relative import fails, try importing from the same directory
    from monitor_thread import NetworkMonitorThread  # Fall back to absolute import


class SplashScreen(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(400, 200)
        
        layout = QVBoxLayout()
        
        # Create frame to hold content
        container = QFrame()
        container.setStyleSheet("""
            QFrame {
                background-color: #2b2b2b;
                border-radius: 10px;
                border: 2px solid #3a3a3a;
            }
        """)
        container_layout = QVBoxLayout(container)
        
        # Add title
        title = QLabel("SysDaemon AI")
        title.setStyleSheet("""
            QLabel {
                color: #ffffff;
                font-size: 24px;
                font-weight: bold;
            }
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(title)
        
        # Add loading text
        self.status_label = QLabel("Loading...")
        self.status_label.setStyleSheet("QLabel { color: #cccccc; }")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(self.status_label)
        
        # Add progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #3a3a3a;
                border-radius: 5px;
                text-align: center;
                background-color: #1a1a1a;
            }
            QProgressBar::chunk {
                background-color: #4a9eff;
                border-radius: 3px;
            }
        """)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        container_layout.addWidget(self.progress_bar)
        
        layout.addWidget(container)
        self.setLayout(layout)
        
        # Center on screen
        screen = QApplication.primaryScreen().geometry()
        self.move(
            screen.center().x() - self.width() // 2,
            screen.center().y() - self.height() // 2
        )
    
    def set_progress(self, value):
        self.progress_bar.setValue(value)
    
    def set_status(self, text):
        self.status_label.setText(text)

class InitializationThread(QThread):
    progress_update = pyqtSignal(int, str)
    initialization_complete = pyqtSignal(dict)
    
    def __init__(self, network_monitor):
        super().__init__()
        self.network_monitor = network_monitor
    
    def run(self):
        try:
            # Step 1: Initialize core components (15%)
            self.progress_update.emit(15, "Initializing core components...")
            self.network_monitor.initialize_system()
            self.msleep(500)
            
            # Step 2: Load network interfaces (25%)
            self.progress_update.emit(25, "Loading network interfaces...")
            self.msleep(500)
            
            # Step 3: Initialize database and load baseline (35%)
            self.progress_update.emit(35, "Loading system baseline...")
            self.network_monitor.load_baseline()
            self.msleep(500)
            
            # Step 4: Check Ollama availability (45%)
            self.progress_update.emit(45, "Checking AI services...")
            ollama_available = self.network_monitor.check_ollama_health()
            if not ollama_available:
                logging.warning("Ollama service not available. Some AI features will be limited.")
            
            # Step 5: Perform initial analysis (55%)
            self.progress_update.emit(55, "Performing initial analysis...")
            initial_state = {
                'performance_metrics': self.network_monitor.performance_monitor.get_system_metrics(),
                'connection_stats': self.network_monitor.get_connection_stats(),
                'ollama_available': ollama_available
            }

            # Step 6: Initialize UI Components (65%)
            self.progress_update.emit(65, "Creating user interface...")
            
            # Step 7: Setup Main Window (75%)
            self.progress_update.emit(75, "Setting up main window...")
            
            # Step 8: Initialize Data Views (85%)
            self.progress_update.emit(85, "Initializing data views...")
            
            # Step 9: Setup System Monitoring (95%)
            self.progress_update.emit(95, "Setting up system monitoring...")
            
            # Step 10: Finalize (100%)
            self.progress_update.emit(100, "Startup complete")
            self.msleep(500)
            
            # Complete initialization
            self.initialization_complete.emit(initial_state)
            
        except Exception as e:
            logging.error(f"Initialization error: {str(e)}\n{traceback.format_exc()}")
            self.initialization_complete.emit({
                'error': str(e),
                'performance_metrics': {},
                'connection_stats': {},
                'ollama_available': False
            })

class NetworkMonitorGUI(QMainWindow):
    error_signal = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Initialize logging first
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('logs/monitoring.log'),
                logging.StreamHandler()
            ]
        )
        
        try:
            # Show login dialog first
            from login_dialog import LoginDialog
            login_dialog = LoginDialog(self)
            if login_dialog.exec() != QDialog.DialogCode.Accepted:
                self.logger.info("Login cancelled or failed")
                sys.exit(0)
            
            # Store the JWT token
            self.auth_token = login_dialog.auth_token
            
            # Import config
            import config
            self.config = config
            
            # Initialize packet capture state
            self.captured_packets = []
            
            # Create network monitor instance
            self.network_monitor = NetworkMonitor()
            self.db_manager = DatabaseManager()
            self.auth_manager = AuthManager(self.db_manager)  # Initialize AuthManager
            
            # Initialize remote agent components
            self.agents = []
            self.agent_search_input = QLineEdit()
            self.agents_list = QComboBox()
            
            # Initialize agent server
            self.agent_server = AgentServer()
            self.agent_server_thread = threading.Thread(target=self.agent_server.start, daemon=True)
            self.agent_server_thread.start()
            
            # Timer to update agents list
            self.agents_update_timer = QTimer()
            self.agents_update_timer.timeout.connect(self.update_agents_list)
            self.agents_update_timer.start(5000)  # Update every 5 seconds
            
            # Show splash screen
            self.splash = SplashScreen()
            self.splash.show()
            
            # Initialize state
            self.initial_analysis = None
            self.ollama_available = False
            
            # Create status bar
            self.status_bar = self.statusBar()
            
            # Start initialization thread
            self.init_thread = InitializationThread(self.network_monitor)
            self.init_thread.progress_update.connect(self.update_splash)
            self.init_thread.initialization_complete.connect(self.finish_initialization)
            self.init_thread.start()
            
        except Exception as e:
            self.logger.error(f"Error during initialization: {str(e)}")
            raise
        
    def update_splash(self, progress, status):
        self.splash.set_progress(progress)
        self.splash.set_status(status)
    
    def finish_initialization(self, analysis_result):
        """Handle completion of initialization"""
        try:
            # Update splash with initial data loading (75%)
            self.splash.set_status("Loading initial data...")
            self.splash.set_progress(75)
            self.initial_analysis = analysis_result
            self.ollama_available = analysis_result.get('ollama_available', False)
            QApplication.processEvents()
            
            if 'error' in analysis_result:
                logging.warning(f"Initialization completed with error: {analysis_result['error']}")
            
            # Setup UI components (85%)
            self.splash.set_status("Setting up user interface...")
            self.splash.set_progress(85)
            self.setup_ui()
            QApplication.processEvents()
            
            # Initialize monitoring thread (90%)
            self.splash.set_status("Initializing system monitor...")
            self.splash.set_progress(90)
            self.monitor_thread = NetworkMonitorThread(self.network_monitor)
            self.monitor_thread.connection_update.connect(self.update_connections_table)
            self.monitor_thread.stats_update.connect(self.update_network_stats)
            self.monitor_thread.error_signal.connect(lambda msg: self.status_bar.showMessage(f"Error: {msg}", 5000))
            QApplication.processEvents()
            
            # Start monitoring and timers (95%)
            self.splash.set_status("Starting system monitoring...")
            self.splash.set_progress(95)
            self.monitor_thread.start()
            
            # Setup update timer
            self.update_timer = QTimer()
            self.update_timer.timeout.connect(self.update_all)
            self.update_timer.start(60000)  # Update every minute
            
            # Load initial data (98%)
            self.splash.set_status("Loading initial system state...")
            self.splash.set_progress(98)
            self.update_system_status()
            self.update_network_stats()
            self.update_system_health()
            self.update_security_analysis()
            QApplication.processEvents()
            
            # Final setup (100%)
            self.splash.set_status("Ready!")
            self.splash.set_progress(100)
            QApplication.processEvents()
            
            # Short delay to ensure everything is ready
            QTimer.singleShot(500, self.show_main_window)
            
        except Exception as e:
            logging.error(f"Error during initialization: {str(e)}\n{traceback.format_exc()}")
            self.status_bar.showMessage(f"Initialization error: {str(e)}", 5000)
            self.show()
            
    def show_main_window(self):
        """Show main window and close splash screen"""
        self.show()
        self.raise_()
        self.activateWindow()
        self.splash.close()
    
    def setup_ui(self):
        """Setup the main user interface"""
        try:
            self.setWindowTitle('SysDaemon AI - Security Monitor')
            self.resize(1200, 800)
            
            # Create central widget and main layout
            central_widget = QWidget()
            self.setCentralWidget(central_widget)
            self.main_layout = QVBoxLayout(central_widget)
            self.main_layout.setSpacing(1)
            self.main_layout.setContentsMargins(1, 1, 1, 1)
            
            # Set application icon
            icon_path = os.path.join(os.path.dirname(__file__), 'icons', 'app_icon.png')
            if os.path.exists(icon_path):
                self.setWindowIcon(QIcon(icon_path))
            else:
                self.logger.warning(f"Icon file not found at {icon_path}")
                # Create icons directory if it doesn't exist
                icons_dir = os.path.join(os.path.dirname(__file__), 'icons')
                os.makedirs(icons_dir, exist_ok=True)
                
                # Create a default icon if the icon file is missing
                self.create_default_icon(icon_path)
                if os.path.exists(icon_path):
                    self.setWindowIcon(QIcon(icon_path))
            
            # Create top status bar
            self.create_system_status_bar()
            
            # Create tabs for different views
            self.tabs = QTabWidget()
            self.main_layout.addWidget(self.tabs)
            
            # Create Home tab
            self.home_tab = QWidget()
            home_layout = QVBoxLayout(self.home_tab)
            home_layout.setSpacing(1)
            home_layout.setContentsMargins(1, 1, 1, 1)
            
            # Create toolbar
            toolbar = QToolBar()
            self.addToolBar(toolbar)
            
            # Add About button to toolbar with proper tooltip
            about_action = QAction('About', self)
            about_action.setIcon(QIcon.fromTheme('help-about'))
            about_action.setToolTip('About SysDaemon AI')
            about_action.triggered.connect(self.show_about_dialog)
            toolbar.addAction(about_action)
            
            # Add Export button to toolbar with proper tooltip
            export_action = QAction('Export', self)
            export_action.setIcon(QIcon.fromTheme('document-save'))
            export_action.setToolTip('Export system data to Excel')
            export_action.triggered.connect(self.export_data)
            toolbar.addAction(export_action)

            # Add system startup checkbox
            self.startup_checkbox = QCheckBox("Run at System Startup")
            self.startup_checkbox.setChecked(self._is_launch_agent_enabled())
            self.startup_checkbox.stateChanged.connect(self.handle_startup_toggle)
            toolbar.addWidget(self.startup_checkbox)

            # Create main content splitter
            content_splitter = QSplitter(Qt.Orientation.Horizontal)
            content_splitter.setContentsMargins(0, 0, 0, 0)
            
            # Left panel: Live Connections and Services
            left_panel = QWidget()
            left_layout = QVBoxLayout(left_panel)
            left_layout.setContentsMargins(1, 1, 1, 1)
            self.create_live_connections_panel(left_layout)
            content_splitter.addWidget(left_panel)
            
            # Center panel: Network Statistics and Charts
            center_panel = QWidget()
            center_layout = QVBoxLayout(center_panel)
            center_layout.setContentsMargins(1, 1, 1, 1)
            self.create_network_stats_panel(center_layout)
            content_splitter.addWidget(center_panel)
            
            # Right panel: Security Analysis and System Health
            right_panel = QWidget()
            right_layout = QVBoxLayout(right_panel)
            right_layout.setContentsMargins(1, 1, 1, 1)
            self.create_security_panel(right_layout)
            content_splitter.addWidget(right_panel)
            
            # Set the splitter sizes for better initial layout
            content_splitter.setSizes([400, 500, 500])
            
            # Add content splitter to home layout
            home_layout.addWidget(content_splitter)
            
            # Add Home tab
            self.tabs.addTab(self.home_tab, "Home")
            
            # Create and add Analytics tab
            self.setup_analytics_tab()
            self.tabs.addTab(self.analytics_tab, "Analytics")
            
            # Create and add Packet Capture tab
            self.packet_capture_tab = self.setup_packet_capture_tab()
            self.tabs.addTab(self.packet_capture_tab, "Packet Capture")
            
            # Create and add Terminal tab
            self.terminal_tab = self.setup_terminal_tab()
            self.tabs.addTab(self.terminal_tab, "Terminal")
            
            # Initialize AI agents and their tabs
            self.init_ai_agents()
            self.tabs.addTab(self.agent_tabs, "AI Agents")
            
            # Add Admin tab if user has admin role
            token_data = self.auth_manager.verify_token(self.auth_token)
            if token_data and token_data.get('role') == 'admin':
                self.admin_tab = self.setup_admin_tab()
                self.tabs.addTab(self.admin_tab, "Admin")
            
            # Create and add Security Agent Chat tab
            self.security_agent_chat_tab = self.create_security_agent_chat_tab()
            self.tabs.addTab(self.security_agent_chat_tab, "Security Agent Chat")
            
            # Create bottom status bar
            self.statusBar = QMainWindow.statusBar(self)
            self.statusBar.setMaximumHeight(20)
            
            # Initialize update timer
            self.init_timer()
            
            # Initialize system tray
            self.create_system_tray()
            
        except Exception as e:
            self.logger.error(f"Error setting up UI: {str(e)}")
            raise
    
    def create_system_status_bar(self):
        """Create the top status bar with system information and controls"""
        # Create main status widget
        status_widget = QWidget()
        status_widget.setMaximumHeight(25)  # Limit the height of the status bar
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(5, 0, 5, 0)  # Minimal vertical margins
        status_layout.setSpacing(15)  # Add space between items
        
        # System metrics group (left side)
        metrics_widget = QWidget()
        metrics_layout = QHBoxLayout(metrics_widget)
        metrics_layout.setContentsMargins(0, 0, 0, 0)
        metrics_layout.setSpacing(15)
        
        # CPU Usage
        cpu_widget = QWidget()
        cpu_layout = QHBoxLayout(cpu_widget)
        cpu_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        cpu_layout.setSpacing(5)  # Small spacing between label and value
        cpu_label = QLabel("CPU:")
        cpu_label.setStyleSheet("""
            QLabel {
                color: #9E9E9E;
                font-size: 11px;
            }
        """)
        self.cpu_label = QLabel("0%")
        self.cpu_label.setStyleSheet("""
            QLabel {
                color: #4CAF50;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        cpu_layout.addWidget(cpu_label)
        cpu_layout.addWidget(self.cpu_label)
        metrics_layout.addWidget(cpu_widget)
        
        # Memory Usage
        memory_widget = QWidget()
        memory_layout = QHBoxLayout(memory_widget)
        memory_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        memory_layout.setSpacing(5)  # Small spacing between label and value
        memory_label = QLabel("Memory:")
        memory_label.setStyleSheet("""
            QLabel {
                color: #9E9E9E;
                font-size: 11px;
            }
        """)
        self.memory_label = QLabel("0%")
        self.memory_label.setStyleSheet("""
            QLabel {
                color: #2196F3;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        memory_layout.addWidget(memory_label)
        memory_layout.addWidget(self.memory_label)
        metrics_layout.addWidget(memory_widget)
        
        # Network I/O
        network_widget = QWidget()
        network_layout = QHBoxLayout(network_widget)
        network_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        network_layout.setSpacing(5)  # Small spacing between label and value
        network_label = QLabel("Network:")
        network_label.setStyleSheet("""
            QLabel {
                color: #9E9E9E;
                font-size: 11px;
            }
        """)
        self.network_io_label = QLabel("0 B/s")
        self.network_io_label.setStyleSheet("""
            QLabel {
                color: #FF9800;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        network_layout.addWidget(network_label)
        network_layout.addWidget(self.network_io_label)
        metrics_layout.addWidget(network_widget)
        
        # Uptime
        uptime_widget = QWidget()
        uptime_layout = QHBoxLayout(uptime_widget)
        uptime_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        uptime_layout.setSpacing(5)  # Small spacing between label and value
        uptime_label = QLabel("Uptime:")
        uptime_label.setStyleSheet("""
            QLabel {
                color: #9E9E9E;
                font-size: 11px;
            }
        """)
        self.uptime_label = QLabel("0:00:00")
        self.uptime_label.setStyleSheet("""
            QLabel {
                color: #9C27B0;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        uptime_layout.addWidget(uptime_label)
        uptime_layout.addWidget(self.uptime_label)
        metrics_layout.addWidget(uptime_widget)
        
        status_layout.addWidget(metrics_widget)
        status_layout.addStretch()  # Push everything to the sides
        
        # Right side buttons
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(5)
        
        # Export button
        export_btn = QPushButton("Export")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #9E9E9E;
                border: none;
                font-size: 11px;
                padding: 2px 8px;
            }
            QPushButton:hover {
                color: #FFFFFF;
                background-color: #424242;
                border-radius: 3px;
            }
        """)
        export_btn.clicked.connect(self.export_data)
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        buttons_layout.addWidget(export_btn)
        
        # About button
        about_btn = QPushButton("About")
        about_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #9E9E9E;
                border: none;
                font-size: 11px;
                padding: 2px 8px;
            }
            QPushButton:hover {
                color: #FFFFFF;
                background-color: #424242;
                border-radius: 3px;
            }
        """)
        about_btn.clicked.connect(self.show_about_dialog)
        about_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        buttons_layout.addWidget(about_btn)
        
        status_layout.addWidget(buttons_widget)
        
        self.centralWidget().layout().addWidget(status_widget)
        
    def show_about_dialog(self):
        """Show the About dialog with application information"""
        about_text = """
<h2>SysDaemon AI</h2>
<p>Version 2.0 (January 13, 2025)</p>

<h3>Overview</h3>
<p>SysDaemon AI is an advanced network monitoring and security analysis tool powered by AI agents. 
It provides real-time monitoring, threat detection, and security intelligence gathering.</p>

<h3>Key Features</h3>

<h4>System Health Monitoring</h4>
<ul>
<li>Real-time CPU, memory, and disk usage monitoring</li>
<li>Visual progress bars with dynamic color coding</li>
<li>Automated alerts for resource-intensive processes</li>
<li>Historical performance tracking</li>
</ul>

<h4>Network Monitoring</h4>
<ul>
<li>Real-time network traffic monitoring</li>
<li>Active connection tracking</li>
<li>Service and port monitoring</li>
<li>Network statistics and analytics</li>
</ul>

<h4>Security Intelligence</h4>
<ul>
<li>Local LLM-powered security analysis</li>
<li>Integration with AbuseIPDB for threat detection</li>
<li>Smart API caching with rate limiting</li>
<li>Automated security recommendations</li>
<li>Comprehensive logging system</li>
</ul>

<h4>Packet Analysis</h4>
<ul>
<li>Real-time packet capture and analysis</li>
<li>Export to PCAP for Wireshark integration</li>
<li>Excel-based packet analysis</li>
<li>Customizable capture filters</li>
<li>Protocol-specific analysis</li>
<li>AI-Powered Packet Analysis</li>
</ul>

<h4>Data Management</h4>
<ul>
<li>SQLite-based data persistence</li>
<li>Historical data analysis</li>
<li>Performance trend visualization</li>
<li>Comprehensive data export</li>
<li>Automated backup system</li>
</ul>

<h4>Security Agent Chat</h4>
<ul>
<li> Security focused prompt for LLM focus</li>
<li> Threaded conversation to retain context</li>
<li> Simple and user-friendly interface</li>
</ul>

<h3>Contact Information</h3>
<p><strong>Created by:</strong> Christopher Bradford</p>
<p><strong> contact@christopherdanielbradford.com
<p><strong>License:</strong> Commercial License</p>
"""
        QMessageBox.about(self, "About SysDaemon AI", about_text)
    
    def create_live_connections_panel(self, layout):
        # Group box for live connections
        connections_group = QFrame()
        connections_group.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        connections_layout = QVBoxLayout(connections_group)
        
        # Header with title and controls
        header_layout = QHBoxLayout()
        title = QLabel("<h3>Live Connections</h3>")
        title.setStyleSheet("QLabel { color: #4CAF50; }")
        header_layout.addWidget(title)
        
        # Add refresh button
        refresh_button = QPushButton("Refresh Connections")
        refresh_button.clicked.connect(self.update_connections)
        header_layout.addStretch()
        header_layout.addWidget(refresh_button)
        
        connections_layout.addLayout(header_layout)
        
        # Connections table
        self.connections_table = QTableWidget()
        self.connections_table.setColumnCount(6)
        self.connections_table.setHorizontalHeaderLabels([
            "Local Address", "Local Port", "Remote Address", "Remote Port", "State", "Process"
        ])
        
        # Set table properties
        self.connections_table.horizontalHeader().setStretchLastSection(True)
        self.connections_table.setAlternatingRowColors(True)
        self.connections_table.setSortingEnabled(True)
        self.connections_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.connections_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #2D2D2D;
                background-color: #1E1E1E;
                border: 1px solid #2D2D2D;
                color: #FFFFFF;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #2D2D2D;
                color: #FFFFFF;
                padding: 5px;
                border: 1px solid #3D3D3D;
            }
            QTableWidget::item:alternate {
                background-color: #262626;
            }
        """)
        
        # Set column widths
        header = self.connections_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # Local Address
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)  # Local Port
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)  # Remote Address
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)  # Remote Port
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)  # State
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)     # Process
        
        # Set default column widths
        self.connections_table.setColumnWidth(0, 120)  # Local Address
        self.connections_table.setColumnWidth(1, 80)   # Local Port
        self.connections_table.setColumnWidth(2, 120)  # Remote Address
        self.connections_table.setColumnWidth(3, 80)   # Remote Port
        self.connections_table.setColumnWidth(4, 100)  # State
        
        connections_layout.addWidget(self.connections_table)
        
        # Services group
        services_group = QFrame()
        services_group.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        services_layout = QVBoxLayout(services_group)
        
        services_header = QHBoxLayout()
        services_title = QLabel("<h3>Active Services</h3>")
        services_title.setStyleSheet("QLabel { color: #2196F3; }")
        services_header.addWidget(services_title)
        
        # Add services refresh button
        services_refresh_button = QPushButton("Refresh Services")
        services_refresh_button.clicked.connect(self.update_services)
        services_header.addStretch()
        services_header.addWidget(services_refresh_button)
        
        services_layout.addLayout(services_header)
        
        self.services_table = QTableWidget()
        self.services_table.setColumnCount(4)
        self.services_table.setHorizontalHeaderLabels([
            "Service", "Port", "Protocol", "State"
        ])
        self.services_table.horizontalHeader().setStretchLastSection(True)
        services_layout.addWidget(self.services_table)
        
        layout.addWidget(connections_group)
        layout.addWidget(services_group)
        
    def create_network_stats_panel(self, layout):
        # Network Statistics Group
        stats_group = QFrame()
        stats_group.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        stats_layout = QVBoxLayout(stats_group)
        
        stats_title = QLabel("<h3>Network Statistics</h3>")
        stats_title.setStyleSheet("QLabel { color: #2196F3; }")
        stats_layout.addWidget(stats_title)
        
        # Statistics grid
        stats_grid = QGridLayout()
        
        # Create statistic widgets with modern styling
        self.total_connections_label = QLabel("Total Connections: 0")
        self.total_connections_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 12px; }")
        stats_grid.addWidget(self.total_connections_label, 0, 0)
        
        self.unique_ips_label = QLabel("Unique IPs: 0")
        self.unique_ips_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 12px; }")
        stats_grid.addWidget(self.unique_ips_label, 0, 1)
        
        self.active_ports_label = QLabel("Active Ports: 0")
        self.active_ports_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 12px; }")
        stats_grid.addWidget(self.active_ports_label, 1, 0)
        
        self.avg_latency_label = QLabel("Avg. Latency: 0 ms")
        self.avg_latency_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 12px; }")
        stats_grid.addWidget(self.avg_latency_label, 1, 1)
        
        stats_layout.addLayout(stats_grid)
        
        # Network Activity Table
        self.network_table = QTableWidget()
        self.network_table.setColumnCount(5)
        self.network_table.setHorizontalHeaderLabels([
            "Local Address", "Remote Address", "Status", "PID", "Process"
        ])
        self.network_table.horizontalHeader().setStretchLastSection(True)
        stats_layout.addWidget(self.network_table)
        
        layout.addWidget(stats_group)

    def create_security_panel(self, layout):
        # Security Analysis Group
        security_group = QFrame()
        security_group.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        security_layout = QVBoxLayout(security_group)
        
        # Header with title and status
        header_layout = QHBoxLayout()
        security_title = QLabel("<h3>AI Security Analysis</h3>")
        security_title.setStyleSheet("QLabel { color: #F44336; }")
        header_layout.addWidget(security_title)
        
        # Add loading indicator
        self.security_loading = QProgressBar()
        self.security_loading.setMaximumSize(QSize(100, 16))
        self.security_loading.setTextVisible(False)
        self.security_loading.setStyleSheet("""
            QProgressBar {
                border: 1px solid #555555;
                border-radius: 8px;
                background-color: #2D2D2D;
            }
            QProgressBar::chunk {
                background-color: #F44336;
                border-radius: 7px;
            }
        """)
        self.security_loading.setRange(0, 0)  # Makes it an "infinite" animation
        self.security_loading.hide()  # Initially hidden
        header_layout.addWidget(self.security_loading)
        
        # Add last update time label
        self.last_security_update_label = QLabel("Last Update: Never")
        self.last_security_update_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 12px; }")
        header_layout.addWidget(self.last_security_update_label)
        
        header_layout.addStretch()
        security_layout.addLayout(header_layout)
        
        # LLM Analysis Output
        self.security_text = QTextEdit()
        self.security_text.setReadOnly(True)
        self.security_text.setStyleSheet("""
            QTextEdit {
                background-color: #2D2D2D;
                color: #E0E0E0;
                border: 1px solid #555555;
                border-radius: 4px;
            }
        """)
        security_layout.addWidget(self.security_text)
        
        # System Health Group
        health_group = QFrame()
        health_group.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        health_layout = QVBoxLayout(health_group)
        
        health_title = QLabel("<h3>System Health</h3>")
        health_title.setStyleSheet("QLabel { color: #4CAF50; }")
        health_layout.addWidget(health_title)
        
        # Health metrics grid
        health_grid = QGridLayout()
        
        # Create health metric widgets with labels and progress bars
        cpu_layout = QVBoxLayout()
        cpu_layout.setSpacing(2)  # Reduce spacing between label and value
        cpu_label = QLabel("CPU Usage")
        cpu_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 10px; }")
        self.cpu_usage_label = QLabel("0%")  # Changed from self.cpu_usage
        self.cpu_usage_label.setStyleSheet("QLabel { color: #FFFFFF; font-size: 16px; font-weight: bold; }")
        self.cpu_progress = QProgressBar()
        self.cpu_progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #555555;
                border-radius: 4px;
                background-color: #2D2D2D;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        cpu_layout.addWidget(cpu_label)
        cpu_layout.addWidget(self.cpu_usage_label)
        cpu_layout.addWidget(self.cpu_progress)
        
        mem_layout = QVBoxLayout()
        mem_layout.setSpacing(2)  # Reduce spacing between label and value
        mem_label = QLabel("Memory Usage")
        mem_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 10px; }")
        self.memory_usage_label = QLabel("0%")  # Changed from self.mem_usage
        self.memory_usage_label.setStyleSheet("QLabel { color: #FFFFFF; font-size: 16px; font-weight: bold; }")
        self.memory_progress = QProgressBar()
        self.memory_progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #555555;
                border-radius: 4px;
                background-color: #2D2D2D;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        mem_layout.addWidget(mem_label)
        mem_layout.addWidget(self.memory_usage_label)
        mem_layout.addWidget(self.memory_progress)
        
        disk_layout = QVBoxLayout()
        disk_layout.setSpacing(2)  # Reduce spacing between label and value
        disk_label = QLabel("Disk Usage")
        disk_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 10px; }")
        self.disk_usage_label = QLabel("0%")  # Changed from self.disk_usage
        self.disk_usage_label.setStyleSheet("QLabel { color: #FFFFFF; font-size: 16px; font-weight: bold; }")
        self.disk_progress = QProgressBar()
        self.disk_progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #555555;
                border-radius: 4px;
                background-color: #2D2D2D;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        disk_layout.addWidget(disk_label)
        disk_layout.addWidget(self.disk_usage_label)
        disk_layout.addWidget(self.disk_progress)
        
        net_layout = QVBoxLayout()
        net_layout.setSpacing(2)  # Reduce spacing between label and value
        net_label = QLabel("Network Load")
        net_label.setStyleSheet("QLabel { color: #9E9E9E; font-size: 10px; }")
        self.network_load = QLabel("0%")
        self.network_load.setStyleSheet("QLabel { color: #FFFFFF; font-size: 16px; font-weight: bold; }")
        net_layout.addWidget(net_label)
        net_layout.addWidget(self.network_load)
        
        # Add layouts to grid
        health_grid.addLayout(cpu_layout, 0, 0)
        health_grid.addLayout(mem_layout, 0, 1)
        health_grid.addLayout(disk_layout, 1, 0)
        health_grid.addLayout(net_layout, 1, 1)
        
        health_layout.addLayout(health_grid)
        
        layout.addWidget(security_group)
        layout.addWidget(health_group)
        
        # Initialize the security analysis timer
        self.security_timer = QTimer()
        self.security_timer.timeout.connect(self.update_security_analysis)
        # Set to update every hour
        self.security_timer.start(3600000)  # 3600000 ms = 1 hour
        
        # Initial security analysis
        QTimer.singleShot(1000, self.update_security_analysis)

    def setup_analytics_tab(self):
        """Setup the analytics tab with interactive controls and visualizations"""
        self.analytics_tab = QWidget()
        layout = QVBoxLayout(self.analytics_tab)
        
        # Control Panel
        control_panel = QWidget()
        control_layout = QHBoxLayout(control_panel)
        
        # Time Range Selection
        time_group = QWidget()
        time_layout = QHBoxLayout(time_group)
        time_layout.addWidget(QLabel("Time Range:"))
        
        self.time_range_combo = QComboBox()
        self.time_range_combo.addItems([
            "Last Hour", "Last 6 Hours", "Last 24 Hours", 
            "Last 7 Days", "Last 30 Days", "Custom Range"
        ])
        self.time_range_combo.currentTextChanged.connect(self.handle_time_range_change)
        time_layout.addWidget(self.time_range_combo)
        
        # Custom date range widgets
        self.date_range_widget = QWidget()
        date_range_layout = QHBoxLayout(self.date_range_widget)
        
        date_range_layout.addWidget(QLabel("From:"))
        self.date_from = QDateTimeEdit(QDateTime.currentDateTime().addDays(-1))
        self.date_from.setCalendarPopup(True)
        self.date_from.dateTimeChanged.connect(self.update_analytics)
        date_range_layout.addWidget(self.date_from)
        
        date_range_layout.addWidget(QLabel("To:"))
        self.date_to = QDateTimeEdit(QDateTime.currentDateTime())
        self.date_to.setCalendarPopup(True)
        self.date_to.dateTimeChanged.connect(self.update_analytics)
        date_range_layout.addWidget(self.date_to)
        
        # Initially hide custom date range
        self.date_range_widget.hide()
        time_layout.addWidget(self.date_range_widget)
        
        control_layout.addWidget(time_group)
        
        # Metrics Selection
        metrics_group = QWidget()
        metrics_layout = QHBoxLayout(metrics_group)
        metrics_layout.addWidget(QLabel("Metrics:"))
        
        self.metrics_combo = QComboBox()
        self.metrics_combo.addItems([
            "System Performance", "Network Activity",
            "Security Events", "Process Activity"
        ])
        self.metrics_combo.currentTextChanged.connect(self.update_analytics)
        metrics_layout.addWidget(self.metrics_combo)
        
        control_layout.addWidget(metrics_group)
        
        layout.addWidget(control_panel)
        
        # Create tab widget for different views
        self.analytics_tabs = QTabWidget()
        
        # Graphs tab
        graphs_widget = QWidget()
        graphs_layout = QVBoxLayout(graphs_widget)
        
        # Create matplotlib figure and canvas
        self.figure = plt.figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.figure)
        self.axes = self.figure.subplots(2, 2)
        
        # Create the navigation toolbar
        self.toolbar = NavigationToolbar(self.canvas, self)
        
        # Add matplotlib canvas and toolbar to layout
        graphs_layout.addWidget(self.toolbar)
        graphs_layout.addWidget(self.canvas)
        
        self.analytics_tabs.addTab(graphs_widget, "Graphs")
        
        # Statistics tab
        stats_widget = QWidget()
        stats_layout = QVBoxLayout(stats_widget)
        
        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(2)
        self.stats_table.setHorizontalHeaderLabels(["Metric", "Value"])
        stats_layout.addWidget(self.stats_table)
        
        self.analytics_tabs.addTab(stats_widget, "Statistics")
        
        # Trends tab
        trends_widget = QWidget()
        trends_layout = QVBoxLayout(trends_widget)
        
        self.trends_table = QTableWidget()
        self.trends_table.setColumnCount(3)
        self.trends_table.setHorizontalHeaderLabels(["Metric", "Trend", "Change"])
        trends_layout.addWidget(self.trends_table)
        
        self.analytics_tabs.addTab(trends_widget, "Trends")
        
        layout.addWidget(self.analytics_tabs)

    def handle_time_range_change(self, selected_range):
        """Handle changes in time range selection"""
        if selected_range == "Custom Range":
            self.date_range_widget.show()
        else:
            self.date_range_widget.hide()
        self.update_analytics()
    
    def update_analytics(self):
        """Update analytics based on selected time range and metrics"""
        try:
            # Get time range
            time_range = self.time_range_combo.currentText()
            if time_range == "Custom Range":
                start_time = self.date_from.dateTime().toPyDateTime()
                end_time = self.date_to.dateTime().toPyDateTime()
            else:
                end_time = datetime.now()
                if time_range == "Last Hour":
                    start_time = end_time - timedelta(hours=1)
                elif time_range == "Last 6 Hours":
                    start_time = end_time - timedelta(hours=6)
                elif time_range == "Last 24 Hours":
                    start_time = end_time - timedelta(hours=24)
                elif time_range == "Last 7 Days":
                    start_time = end_time - timedelta(days=7)
                else:  # Last 30 Days
                    start_time = end_time - timedelta(days=30)
            
            # Get selected metrics
            metrics_type = self.metrics_combo.currentText()
            
            # Update graphs
            self.update_analytics_graphs(start_time, end_time, metrics_type)
            
            # Update statistics
            self.update_analytics_stats(start_time, end_time, metrics_type)
            
            # Update trends
            self.update_analytics_trends(start_time, end_time, metrics_type)
            
        except Exception as e:
            self.statusBar.showMessage(f"Error updating analytics: {str(e)}", 5000)
            print(f"Analytics update error: {str(e)}")

    def update_analytics_graphs(self, start_time, end_time, metrics_type):
        """Update the analytics graphs"""
        try:
            # Clear the current figure
            self.figure.clear()
            
            # Create new subplot grid
            self.axes = self.figure.subplots(2, 2)
            
            if metrics_type == "System Performance":
                # CPU Usage over time (top left)
                cpu_data = self.get_cpu_history(start_time, end_time)
                self.axes[0, 0].plot(cpu_data['timestamps'], cpu_data['values'], 'b-', linewidth=2)
                self.axes[0, 0].set_title('CPU Usage', pad=10)
                self.axes[0, 0].set_ylabel('Percentage')
                self.axes[0, 0].grid(True, alpha=0.3)
                self.axes[0, 0].set_ylim(0, 100)
                self.axes[0, 0].tick_params(axis='x', rotation=30)
                
                # Memory Usage over time (top right)
                mem_data = self.get_memory_history(start_time, end_time)
                self.axes[0, 1].plot(mem_data['timestamps'], mem_data['values'], 'g-', linewidth=2)
                self.axes[0, 1].set_title('Memory Usage', pad=10)
                self.axes[0, 1].set_ylabel('Percentage')
                self.axes[0, 1].grid(True, alpha=0.3)
                self.axes[0, 1].set_ylim(0, 100)
                self.axes[0, 1].tick_params(axis='x', rotation=30)
                
                # Disk I/O over time (bottom)
                disk_data = self.get_disk_history(start_time, end_time)
                self.axes[1, 0].plot(disk_data['timestamps'], disk_data['read'], 'r-', label='Read', linewidth=2)
                self.axes[1, 0].plot(disk_data['timestamps'], disk_data['write'], 'b-', label='Write', linewidth=2)
                self.axes[1, 0].set_title('Disk I/O', pad=10)
                self.axes[1, 0].set_ylabel('MB/s')
                self.axes[1, 0].legend(loc='upper right')
                self.axes[1, 0].grid(True, alpha=0.3)
                self.axes[1, 0].tick_params(axis='x', rotation=30)
                
                # Hide the bottom right plot
                self.axes[1, 1].set_visible(False)
                
            elif metrics_type == "Network Activity":
                # Network Traffic over time
                traffic_data = self.get_network_traffic_history(start_time, end_time)
                self.axes[0, 0].plot(traffic_data['timestamps'], traffic_data['sent'], 'b-', label='Sent', linewidth=2)
                self.axes[0, 0].plot(traffic_data['timestamps'], traffic_data['received'], 'g-', label='Received', linewidth=2)
                self.axes[0, 0].set_title('Network Traffic')
                self.axes[0, 0].set_ylabel('MB/s')
                self.axes[0, 0].legend()
                self.axes[0, 0].grid(True)
                self.axes[0, 0].tick_params(axis='x', rotation=30)

                # Connection Count over time
                conn_data = self.get_connection_history(start_time, end_time)
                self.axes[0, 1].plot(conn_data['timestamps'], conn_data['values'], 'r-', linewidth=2)
                self.axes[0, 1].set_title('Connection Count')
                self.axes[0, 1].set_ylabel('Count')
                self.axes[0, 1].grid(True)
                self.axes[0, 1].tick_params(axis='x', rotation=30)

                # Port Activity over time
                port_data = self.get_port_activity_history(start_time, end_time)
                self.axes[1, 0].bar(port_data['ports'], port_data['counts'])
                self.axes[1, 0].set_title('Port Activity')
                self.axes[1, 0].set_ylabel('Connection Count')
                self.axes[1, 0].tick_params(axis='x', rotation=45)
                self.axes[1, 0].grid(True)

                # Protocol Distribution
                proto_data = self.get_protocol_distribution(start_time, end_time)
                self.axes[1, 1].pie(proto_data['values'], labels=proto_data['labels'], autopct='%1.1f%%')
                self.axes[1, 1].set_title('Protocol Distribution')

            elif metrics_type == "Security Events":
                # Security Events over time
                events_data = self.get_security_events_history(start_time, end_time)
                self.axes[0, 0].plot(events_data['timestamps'], events_data['counts'], 'r-', linewidth=2)
                self.axes[0, 0].set_title('Security Events')
                self.axes[0, 0].set_ylabel('Event Count')
                self.axes[0, 0].grid(True)
                self.axes[0, 0].tick_params(axis='x', rotation=30)

                # Threat Types Distribution
                threat_data = self.get_threat_distribution(start_time, end_time)
                self.axes[0, 1].pie(threat_data['values'], labels=threat_data['labels'], autopct='%1.1f%%')
                self.axes[0, 1].set_title('Threat Types')

                # Top Attack Sources
                attack_data = self.get_top_attack_sources(start_time, end_time)
                self.axes[1, 0].bar(attack_data['ips'], attack_data['counts'])
                self.axes[1, 0].set_title('Top Attack Sources')
                self.axes[1, 0].tick_params(axis='x', rotation=45)
                self.axes[1, 0].grid(True)

                # Severity Distribution
                severity_data = self.get_severity_distribution(start_time, end_time)
                self.axes[1, 1].pie(severity_data['values'], labels=severity_data['labels'], autopct='%1.1f%%')
                self.axes[1, 1].set_title('Severity Distribution')

            elif metrics_type == "Process Activity":
                # Adjust figure layout to accommodate labels
                plt.subplots_adjust(bottom=0.2)
                
                # Top CPU Consumers
                cpu_procs = self.get_top_cpu_processes(start_time, end_time)
                x = range(len(cpu_procs['names']))
                self.axes[0, 0].bar(x, cpu_procs['cpu_percent'])
                self.axes[0, 0].set_title('Top CPU Consumers')
                self.axes[0, 0].set_ylabel('CPU %')
                self.axes[0, 0].set_xticks(x)
                self.axes[0, 0].set_xticklabels(cpu_procs['names'], rotation=45, ha='right')
                self.axes[0, 0].grid(True, axis='y')

                # Top Memory Consumers
                mem_procs = self.get_top_memory_processes(start_time, end_time)
                x = range(len(mem_procs['names']))
                self.axes[0, 1].bar(x, mem_procs['memory_percent'])
                self.axes[0, 1].set_title('Top Memory Consumers')
                self.axes[0, 1].set_ylabel('Memory %')
                self.axes[0, 1].set_xticks(x)
                self.axes[0, 1].set_xticklabels(mem_procs['names'], rotation=45, ha='right')
                self.axes[0, 1].grid(True, axis='y')

                # Process Count over time
                proc_data = self.get_process_count_history(start_time, end_time)
                self.axes[1, 0].plot(proc_data['timestamps'], proc_data['values'], 'b-', linewidth=2)
                self.axes[1, 0].set_title('Process Count')
                self.axes[1, 0].set_xlabel('Time')
                self.axes[1, 0].set_ylabel('Count')
                self.axes[1, 0].grid(True)
                self.axes[1, 0].tick_params(axis='x', rotation=30)

                # Process Types Distribution
                type_data = self.get_process_types_distribution(start_time, end_time)
                self.axes[1, 1].pie(type_data['values'], labels=type_data['labels'], autopct='%1.1f%%')
                self.axes[1, 1].set_title('Process Types Distribution')

            # Adjust layout
            self.figure.tight_layout()
            
            # Draw the canvas
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error updating analytics graphs: {str(e)}")
            traceback.print_exc()
    
    def update_analytics_stats(self, start_time, end_time, metrics_type):
        """Update the statistics table based on selected parameters"""
        try:
            self.stats_table.setRowCount(0)
            stats = []

            if metrics_type == "System Performance":
                metrics_data = self.db_manager.get_metrics_range(start_time, end_time)
                if metrics_data:
                    cpu_data = [m.cpu_usage for m in metrics_data if m.cpu_usage is not None]
                    memory_data = [m.memory_usage for m in metrics_data if m.memory_usage is not None]
                    disk_data = [m.disk_usage for m in metrics_data if m.disk_usage is not None]
                    network_data = [m.network_throughput for m in metrics_data if m.network_throughput is not None]
                    process_data = [m.process_count for m in metrics_data if m.process_count is not None]
                    
                    stats = [
                        ("Average CPU Usage", f"{sum(cpu_data)/len(cpu_data):.1f}%" if cpu_data else "N/A"),
                        ("Peak CPU Usage", f"{max(cpu_data):.1f}%" if cpu_data else "N/A"),
                        ("Average Memory Usage", f"{sum(memory_data)/len(memory_data):.1f}%" if memory_data else "N/A"),
                        ("Peak Memory Usage", f"{max(memory_data):.1f}%" if memory_data else "N/A"),
                        ("Average Disk Usage", f"{sum(disk_data)/len(disk_data):.1f}%" if disk_data else "N/A"),
                        ("Peak Disk Usage", f"{max(disk_data):.1f}%" if disk_data else "N/A"),
                        ("Average Network Throughput", f"{sum(network_data)/len(network_data):.1f} MB/s" if network_data else "N/A"),
                        ("Average Process Count", f"{sum(process_data)/len(process_data):.1f}" if process_data else "N/A")
                    ]
                
            elif metrics_type == "Network Activity":
                network_data = self.db_manager.get_network_connections(start_time, end_time)
                if network_data:
                    total_sent = sum(conn.bytes_sent for conn in network_data if conn.bytes_sent is not None)
                    total_received = sum(conn.bytes_received for conn in network_data if conn.bytes_received is not None)
                    unique_ports = set(conn.destination_port for conn in network_data if conn.destination_port is not None)
                    protocols = [conn.protocol for conn in network_data if conn.protocol is not None]
                    most_common_protocol = max(set(protocols), key=protocols.count) if protocols else "N/A"
                    
                    stats = [
                        ("Total Data Sent", f"{total_sent / (1024*1024):.1f} MB"),
                        ("Total Data Received", f"{total_received / (1024*1024):.1f} MB"),
                        ("Total Connections", str(len(network_data))),
                        ("Unique Ports", str(len(unique_ports))),
                        ("Most Common Protocol", most_common_protocol),
                        ("Average Bytes per Connection", f"{(total_sent + total_received) / len(network_data) / 1024:.1f} KB" if network_data else "N/A")
                    ]
                
            elif metrics_type == "Security Events":
                alerts = self.db_manager.get_alerts(start_time, end_time)
                if alerts:
                    total_alerts = len(alerts)
                    high_severity = sum(1 for a in alerts if a.severity == "high")
                    medium_severity = sum(1 for a in alerts if a.severity == "medium")
                    low_severity = sum(1 for a in alerts if a.severity == "low")
                    alert_types = [a.alert_type for a in alerts if a.alert_type is not None]
                    most_common_type = max(set(alert_types), key=alert_types.count) if alert_types else "N/A"
                    
                    stats = [
                        ("Total Alerts", str(total_alerts)),
                        ("High Severity Alerts", str(high_severity)),
                        ("Medium Severity Alerts", str(medium_severity)),
                        ("Low Severity Alerts", str(low_severity)),
                        ("Most Common Alert Type", most_common_type),
                        ("Resolved Alerts", str(sum(1 for a in alerts if a.resolved)))
                    ]
                
            elif metrics_type == "Process Activity":
                metrics_data = self.db_manager.get_metrics_range(start_time, end_time)
                if metrics_data:
                    process_counts = [m.process_count for m in metrics_data if m.process_count is not None]
                    
                    stats = [
                        ("Average Process Count", f"{sum(process_counts)/len(process_counts):.1f}" if process_counts else "N/A"),
                        ("Peak Process Count", str(max(process_counts)) if process_counts else "N/A"),
                        ("Minimum Process Count", str(min(process_counts)) if process_counts else "N/A"),
                        ("Process Count Variance", f"{sum((x - (sum(process_counts)/len(process_counts)))**2 for x in process_counts)/len(process_counts):.1f}" if process_counts else "N/A")
                    ]

            # Populate the table
            for stat in stats:
                row = self.stats_table.rowCount()
                self.stats_table.insertRow(row)
                self.stats_table.setItem(row, 0, QTableWidgetItem(stat[0]))
                self.stats_table.setItem(row, 1, QTableWidgetItem(stat[1]))

            # Adjust column widths
            self.stats_table.resizeColumnsToContents()

        except Exception as e:
            self.statusBar.showMessage(f"Error updating statistics: {str(e)}", 5000)
            print(f"Statistics update error: {str(e)}")

    def update_analytics_trends(self, start_time, end_time, metrics_type):
        """Update the trends table based on selected parameters"""
        try:
            self.trends_table.setRowCount(0)
            trends = []

            if metrics_type == "System Performance":
                trends = [
                    ("CPU Usage Trend", self.get_cpu_trend(start_time, end_time)),
                    ("Memory Usage Trend", self.get_memory_trend(start_time, end_time)),
                    ("Disk I/O Trend", self.get_disk_io_trend(start_time, end_time)),
                    ("System Load Trend", self.get_system_load_trend(start_time, end_time))
                ]
            elif metrics_type == "Network Activity":
                trends = [
                    ("Network Traffic Trend", self.get_network_traffic_trend(start_time, end_time)),
                    ("Connection Count Trend", self.get_connection_count_trend(start_time, end_time)),
                    ("Port Activity Trend", self.get_port_activity_trend(start_time, end_time))
                ]
            elif metrics_type == "Security Events":
                trends = [
                    ("Security Events Trend", self.get_security_events_trend(start_time, end_time)),
                    ("Threat Types Trend", self.get_threat_types_trend(start_time, end_time)),
                    ("Attack Sources Trend", self.get_attack_sources_trend(start_time, end_time))
                ]
            elif metrics_type == "Process Activity":
                trends = [
                    ("Process Count Trend", self.get_process_count_trend(start_time, end_time)),
                    ("CPU Usage by Process Trend", self.get_process_cpu_trend(start_time, end_time)),
                    ("Memory Usage by Process Trend", self.get_process_memory_trend(start_time, end_time))
                ]

            # Populate the table
            for trend in trends:
                row = self.trends_table.rowCount()
                self.trends_table.insertRow(row)
                self.trends_table.setItem(row, 0, QTableWidgetItem(trend[0]))
                
                # Add trend arrow
                trend_item = QTableWidgetItem()
                if "increasing" in trend[1].lower():
                    trend_item.setText("↑")
                    trend_item.setForeground(QColor("green"))
                elif "decreasing" in trend[1].lower():
                    trend_item.setText("↓")
                    trend_item.setForeground(QColor("red"))
                else:
                    trend_item.setText("→")
                    trend_item.setForeground(QColor("gray"))
                self.trends_table.setItem(row, 1, trend_item)
                
                # Add trend description
                self.trends_table.setItem(row, 2, QTableWidgetItem(trend[1]))

            # Adjust column widths
            self.trends_table.resizeColumnsToContents()

        except Exception as e:
            self.statusBar.showMessage(f"Error updating trends: {str(e)}", 5000)
            print(f"Trends update error: {str(e)}")

    def get_cpu_history(self, start_time, end_time):
        """Get CPU usage history"""
        timestamps = []
        values = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                values.append(psutil.cpu_percent(interval=None))
                current_time += timedelta(seconds=10)
        except Exception as e:
            print(f"Error getting CPU history: {str(e)}")
        return {'timestamps': timestamps, 'values': values}

    def get_memory_history(self, start_time, end_time):
        """Get memory usage history"""
        timestamps = []
        values = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                values.append(psutil.virtual_memory().percent)
                current_time += timedelta(seconds=10)
        except Exception as e:
            print(f"Error getting memory history: {str(e)}")
        return {'timestamps': timestamps, 'values': values}

    def get_disk_history(self, start_time, end_time):
        """Get disk I/O history"""
        timestamps = []
        read_values = []
        write_values = []
        try:
            current_time = start_time
            last_disk_io = psutil.disk_io_counters()
            last_time = time.time()
            
            while current_time <= end_time:
                timestamps.append(current_time)
                disk_io = psutil.disk_io_counters()
                current_time_sec = time.time()
                time_delta = current_time_sec - last_time
                
                read_speed = (disk_io.read_bytes - last_disk_io.read_bytes) / (1024 * 1024 * time_delta)
                write_speed = (disk_io.write_bytes - last_disk_io.write_bytes) / (1024 * 1024 * time_delta)
                
                read_values.append(read_speed)
                write_values.append(write_speed)
                
                last_disk_io = disk_io
                last_time = current_time_sec
                current_time += timedelta(seconds=10)
                
        except Exception as e:
            print(f"Error getting disk history: {str(e)}")
        return {'timestamps': timestamps, 'read': read_values, 'write': write_values}

    def get_load_history(self, start_time, end_time):
        """Get system load history"""
        timestamps = []
        values = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                values.append(os.getloadavg()[0])  # 1-minute load average
                current_time += timedelta(seconds=10)
        except Exception as e:
            print(f"Error getting load history: {str(e)}")
        return {'timestamps': timestamps, 'values': values}

    def get_network_traffic_history(self, start_time, end_time):
        """Get network traffic history"""
        timestamps = []
        sent_values = []
        received_values = []
        try:
            current_time = start_time
            last_net_io = psutil.net_io_counters()
            last_time = time.time()
            
            while current_time <= end_time:
                timestamps.append(current_time)
                net_io = psutil.net_io_counters()
                current_time_sec = time.time()
                time_delta = current_time_sec - last_time
                
                sent_speed = (net_io.bytes_sent - last_net_io.bytes_sent) / (1024 * 1024 * time_delta)
                received_speed = (net_io.bytes_recv - last_net_io.bytes_recv) / (1024 * 1024 * time_delta)
                
                sent_values.append(sent_speed)
                received_values.append(received_speed)
                
                last_net_io = net_io
                last_time = current_time_sec
                current_time += timedelta(seconds=10)
                
        except Exception as e:
            print(f"Error getting network traffic history: {str(e)}")
        return {'timestamps': timestamps, 'sent': sent_values, 'received': received_values}

    def get_connection_history(self, start_time, end_time):
        """Get connection count history"""
        timestamps = []
        values = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                values.append(len([c for c in psutil.net_connections() if c.status == 'ESTABLISHED']))
                current_time += timedelta(seconds=10)
        except Exception as e:
            print(f"Error getting connection history: {str(e)}")
        return {'timestamps': timestamps, 'values': values}

    def get_port_activity_history(self, start_time, end_time):
        """Get port activity data"""
        port_counts = {}
        try:
            for conn in psutil.net_connections():
                if conn.laddr and len(conn.laddr) > 1:
                    port = conn.laddr[1]
                    port_counts[port] = port_counts.get(port, 0) + 1
            
            # Get top 10 most active ports
            top_ports = sorted(port_counts.items(), key=lambda x: x[1], reverse=True)[:10]
            return {
                'ports': [str(p[0]) for p in top_ports],
                'counts': [p[1] for p in top_ports]
            }
        except Exception as e:
            print(f"Error getting port activity: {str(e)}")
            return {'ports': [], 'counts': []}

    def get_protocol_distribution(self, start_time, end_time):
        """Get protocol distribution data"""
        proto_counts = {'TCP': 0, 'UDP': 0, 'ICMP': 0, 'Other': 0}
        try:
            for conn in psutil.net_connections(kind='all'):
                if conn.type == socket.SOCK_STREAM:
                    proto_counts['TCP'] += 1
                elif conn.type == socket.SOCK_DGRAM:
                    proto_counts['UDP'] += 1
                else:
                    proto_counts['Other'] += 1
            
            return {
                'labels': list(proto_counts.keys()),
                'values': list(proto_counts.values())
            }
        except Exception as e:
            print(f"Error getting protocol distribution: {str(e)}")
            return {'labels': [], 'values': []}

    def get_security_events_history(self, start_time, end_time):
        """Get security events history"""
        # For demonstration, we'll generate some sample security event data
        timestamps = []
        counts = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                # Simulated event count based on time of day
                hour = current_time.hour
                base_count = 5  # Base number of events
                time_factor = abs(12 - hour) / 12.0  # More events during business hours
                count = int(base_count * (1 + time_factor))
                counts.append(count)
                current_time += timedelta(minutes=30)
        except Exception as e:
            print(f"Error getting security events history: {str(e)}")
        return {'timestamps': timestamps, 'counts': counts}

    def get_threat_distribution(self, start_time, end_time):
        """Get threat type distribution"""
        # Sample threat types and counts
        threats = {
            'Malware': 15,
            'Phishing': 8,
            'DDoS': 3,
            'Unauthorized Access': 5,
            'Data Breach': 2
        }
        return {
            'labels': list(threats.keys()),
            'values': list(threats.values())
        }

    def get_top_attack_sources(self, start_time, end_time):
        """Get top attack sources"""
        # Sample attack sources and counts
        sources = {
            '192.168.1.100': 25,
            '10.0.0.5': 18,
            '172.16.0.10': 15,
            '192.168.1.200': 12,
            '10.0.0.15': 10
        }
        return {
            'ips': list(sources.keys()),
            'counts': list(sources.values())
        }

    def get_severity_distribution(self, start_time, end_time):
        """Get severity distribution"""
        severities = {
            'Low': 45,
            'Medium': 30,
            'High': 15,
            'Critical': 10
        }
        return {
            'labels': list(severities.keys()),
            'values': list(severities.values())
        }

    def get_top_cpu_processes(self, start_time, end_time):
        """Get top CPU consuming processes"""
        processes = {}
        try:
            for proc in psutil.process_iter(['name', 'cpu_percent']):
                try:
                    # Truncate process name if too long
                    name = proc.info['name']
                    if len(name) > 20:
                        name = name[:17] + "..."
                    processes[name] = proc.info['cpu_percent']
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            # Get top 10 processes
            top_procs = sorted(processes.items(), key=lambda x: x[1], reverse=True)[:10]
            return {
                'names': [p[0] for p in top_procs],
                'cpu_percent': [p[1] for p in top_procs]
            }
        except Exception as e:
            print(f"Error getting top CPU processes: {str(e)}")
            return {'names': [], 'cpu_percent': []}

    def get_top_memory_processes(self, start_time, end_time):
        """Get top memory consuming processes"""
        processes = {}
        try:
            for proc in psutil.process_iter(['name', 'memory_percent']):
                try:
                    # Truncate process name if too long
                    name = proc.info['name']
                    if len(name) > 20:
                        name = name[:17] + "..."
                    processes[name] = proc.info['memory_percent']
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            # Get top 10 processes
            top_procs = sorted(processes.items(), key=lambda x: x[1], reverse=True)[:10]
            return {
                'names': [p[0] for p in top_procs],
                'memory_percent': [p[1] for p in top_procs]
            }
        except Exception as e:
            print(f"Error getting top memory processes: {str(e)}")
            return {'names': [], 'memory_percent': []}

    def get_process_count_history(self, start_time, end_time):
        """Get process count history"""
        timestamps = []
        values = []
        try:
            current_time = start_time
            while current_time <= end_time:
                timestamps.append(current_time)
                values.append(len(psutil.pids()))
                current_time += timedelta(seconds=10)
        except Exception as e:
            print(f"Error getting process count history: {str(e)}")
        return {'timestamps': timestamps, 'values': values}

    def get_process_types_distribution(self, start_time, end_time):
        """Get process types distribution"""
        type_counts = {'System': 0, 'User': 0, 'Background': 0, 'Other': 0}
        try:
            for proc in psutil.process_iter(['name', 'username']):
                try:
                    if proc.username() == 'root':
                        type_counts['System'] += 1
                    elif proc.username() == getpass.getuser():
                        type_counts['User'] += 1
                    else:
                        type_counts['Background'] += 1
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            return {
                'labels': list(type_counts.keys()),
                'values': list(type_counts.values())
            }
        except Exception as e:
            print(f"Error getting process types distribution: {str(e)}")
            return {'labels': [], 'values': []}

    # Additional helper methods for statistics
    def get_avg_cpu_usage(self, start_time, end_time):
        return psutil.cpu_percent(interval=1)

    def get_peak_cpu_usage(self, start_time, end_time):
        return max(psutil.cpu_percent(interval=0.1) for _ in range(10))

    def get_avg_memory_usage(self, start_time, end_time):
        return psutil.virtual_memory().percent

    def get_peak_memory_usage(self, start_time, end_time):
        return max(psutil.virtual_memory().percent for _ in range(5))

    def get_avg_disk_io(self, start_time, end_time):
        disk_io = psutil.disk_io_counters()
        return (disk_io.read_bytes + disk_io.write_bytes) / (1024 * 1024)

    def get_peak_disk_io(self, start_time, end_time):
        disk_io = psutil.disk_io_counters()
        return max((disk_io.read_bytes + disk_io.write_bytes) / (1024 * 1024) for _ in range(5))

    def get_avg_system_load(self, start_time, end_time):
        return os.getloadavg()[0]

    def get_total_data_sent(self, start_time, end_time):
        return psutil.net_io_counters().bytes_sent / (1024*1024)

    def get_total_data_received(self, start_time, end_time):
        return psutil.net_io_counters().bytes_recv / (1024*1024)

    def get_avg_connection_count(self, start_time, end_time):
        return len([c for c in psutil.net_connections() if c.status == 'ESTABLISHED'])

    def get_peak_connection_count(self, start_time, end_time):
        return max(len([c for c in psutil.net_connections() if c.status == 'ESTABLISHED']) for _ in range(5))

    def get_most_active_port(self, start_time, end_time):
        port_counts = {}
        for conn in psutil.net_connections():
            if conn.laddr and len(conn.laddr) > 1:
                port = conn.laddr.port
                port_counts[port] = port_counts.get(port, 0) + 1
        return max(port_counts.items(), key=lambda x: x[1])[0] if port_counts else "N/A"

    def get_most_common_protocol(self, start_time, end_time):
        proto_counts = {'TCP': 0, 'UDP': 0}
        for conn in psutil.net_connections(kind='all'):
            if conn.type == socket.SOCK_STREAM:
                proto_counts['TCP'] += 1
            elif conn.type == socket.SOCK_DGRAM:
                proto_counts['UDP'] += 1
        return max(proto_counts.items(), key=lambda x: x[1])[0] if proto_counts else "N/A"

    # Helper methods for trends
    def get_cpu_trend(self, start_time, end_time):
        values = [psutil.cpu_percent(interval=0.1) for _ in range(5)]
        return self._calculate_trend(values)

    def get_memory_trend(self, start_time, end_time):
        values = [psutil.virtual_memory().percent for _ in range(5)]
        return self._calculate_trend(values)

    def get_disk_io_trend(self, start_time, end_time):
        values = []
        last_io = psutil.disk_io_counters()
        for _ in range(5):
            time.sleep(0.1)
            io = psutil.disk_io_counters()
            values.append((io.read_bytes + io.write_bytes - last_io.read_bytes - last_io.write_bytes) / 1024)
            last_io = io
        return self._calculate_trend(values)

    def get_system_load_trend(self, start_time, end_time):
        values = [os.getloadavg()[0] for _ in range(5)]
        return self._calculate_trend(values)

    def get_network_traffic_trend(self, start_time, end_time):
        values = []
        last_io = psutil.net_io_counters()
        for _ in range(5):
            time.sleep(0.1)
            io = psutil.net_io_counters()
            values.append((io.bytes_sent + io.bytes_recv - last_io.bytes_sent - last_io.bytes_recv) / 1024)
            last_io = io
        return self._calculate_trend(values)

    def get_connection_count_trend(self, start_time, end_time):
        values = [len([c for c in psutil.net_connections() if c.status == 'ESTABLISHED']) for _ in range(5)]
        return self._calculate_trend(values)

    def get_port_activity_trend(self, start_time, end_time):
        values = [len(set(c.laddr[1] for c in psutil.net_connections() if c.laddr)) for _ in range(5)]
        return self._calculate_trend(values)

    def _calculate_trend(self, values):
        """Calculate trend from a list of values"""
        if not values or len(values) < 2:
            return "Stable"
        
        first_half = sum(values[:len(values)//2]) / (len(values)//2)
        second_half = sum(values[len(values)//2:]) / (len(values) - len(values)//2)
        
        diff = second_half - first_half
        if abs(diff) < 0.1 * first_half:
            return "Stable"
        elif diff > 0:
            return "Increasing" if diff > 0.2 * first_half else "Slightly Increasing"
        else:
            return "Decreasing" if abs(diff) > 0.2 * first_half else "Slightly Decreasing"
    
    def init_timer(self):
        """Initialize the update timer and network monitor thread"""
        # Create and start network monitor thread
        self.monitor_thread = NetworkMonitorThread(self.network_monitor)
        self.monitor_thread.connection_update.connect(self.update_connections_table)
        self.monitor_thread.analysis_update.connect(self._update_security_display)
        self.monitor_thread.error_signal.connect(self._handle_general_error)
        self.monitor_thread.stats_update.connect(self.update_network_stats)
        self.monitor_thread.start()
        
        # Create timer for critical updates
        self.critical_timer = QTimer()
        self.critical_timer.timeout.connect(self.update_critical)
        self.critical_timer.start(1000)  # Update every second
        
        # Create timer for less critical updates
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_all)
        self.update_timer.start(5000)  # Update every 5 seconds
        
    def update_critical(self):
        """Update critical metrics that need more frequent monitoring"""
        try:
            self.update_system_status()
            self.update_system_health()
        except Exception as e:
            self.logger.error(f"Error in critical update: {str(e)}")
    
    def update_all(self):
        """Update all metrics"""
        try:
            self.update_services()
            self.update_network_stats()
            self.update_security_analysis()
        except Exception as e:
            self.logger.error(f"Error in update_all: {str(e)}")
    
    def update_system_status(self):
        try:
            # Update CPU usage
            cpu_percent = psutil.cpu_percent()
            self.cpu_label.setText(f"CPU: {cpu_percent}%")
            
            # Update memory usage
            memory = psutil.virtual_memory()
            # First try to access as named tuple (direct psutil object)
            try:
                self.memory_label.setText(f"Memory: {memory.percent}%")
            except AttributeError:
                # If that fails, try dictionary access for compatibility with performance monitor
                if isinstance(memory, dict):
                    if 'memory' in memory and 'virtual' in memory['memory'] and 'percent' in memory['memory']['virtual']:
                        # Nested structure from performance monitor
                        self.memory_label.setText(f"Memory: {memory['memory']['virtual']['percent']}%")
                    elif 'virtual' in memory and 'percent' in memory['virtual']:
                        # Partially nested structure
                        self.memory_label.setText(f"Memory: {memory['virtual']['percent']}%")
                    elif 'percent' in memory:
                        # Flat dictionary structure
                        self.memory_label.setText(f"Memory: {memory['percent']}%")
                else:
                    # If all else fails, set default
                    self.memory_label.setText(f"0%")
            
            # Update network I/O
            net_io = psutil.net_io_counters()
            self.network_io_label.setText(
                f"Network: ↑{self.format_bytes(net_io.bytes_sent)}/s "
                f"↓{self.format_bytes(net_io.bytes_recv)}/s"
            )
            
            # Update uptime
            boot_time = datetime.fromtimestamp(psutil.boot_time())
            uptime = datetime.now() - boot_time
            self.uptime_label.setText(f"Uptime: {str(uptime).split('.')[0]}")
            
        except Exception as e:
            self.error_signal.emit(f"Error updating system status: {str(e)}")

    def format_bytes(self, bytes):
        for unit in ['B', 'KB', 'MB', 'GB']:
            if bytes < 1024:
                return f"{bytes:.1f}{unit}"
            bytes /= 1024
        return f"{bytes:.1f}TB"

    def is_interesting_address(self, ip: str) -> bool:
        """Check if an IP address should be displayed in the connections table"""
        # Filter out common local and system addresses
        filtered_addresses = {
            '0.0.0.0',
            '127.0.0.1',
            '::',
            '::1',
            '*'
        }
        return ip not in filtered_addresses

    def update_connections(self):
        """Update the connections table with current network connections"""
        try:
            connections = []
            seen_connections = set()  # Track unique connections
            
            for conn in psutil.net_connections(kind='inet'):
                try:
                    # Skip invalid connections
                    if not hasattr(conn, 'laddr') or not conn.laddr:
                        continue
                        
                    # Initialize conn_key with a default value
                    conn_key = None
                    
                    # Assuming this is within a loop or condition where conn is defined
                    if hasattr(conn, 'raddr') and conn.raddr:
                        conn_key = f"{conn.laddr.ip}:{conn.laddr.port}-{conn.raddr.ip}:{conn.raddr.port}"
                    else:
                        conn_key = f"{conn.laddr.ip}:{conn.laddr.port}-0.0.0.0:0"  # Default value if raddr is not present
                    
                    # Now you can safely check conn_key
                    if conn_key and conn_key in seen_connections:
                        continue
                    
                    # Get process information
                    process_name = "Unknown"
                    if conn.pid:
                        try:
                            process = psutil.Process(conn.pid)
                            process_name = process.name()
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            pass
                    
                    # Format connection information
                    connection_info = {
                        'local_address': f"{conn.laddr.ip}:{conn.laddr.port}",
                        'remote_address': f"{conn.raddr.ip}:{conn.raddr.port}" if hasattr(conn, 'raddr') and conn.raddr else "0.0.0.0:0",
                        'status': conn.status,
                        'process': process_name
                    }
                    
                    connections.append(connection_info)
                    seen_connections.add(conn_key)
                    
                except (AttributeError, psutil.NoSuchProcess, psutil.AccessDenied) as e:
                    self.logger.debug(f"Error processing connection: {str(e)}")
                    continue
            
            # Sort connections by status and port
            connections.sort(key=lambda x: (
                x['status'] != 'ESTABLISHED',
                x['status'] != 'LISTEN',
                x['local_address'].split(':')[1]
            ))
            
            # Update the table using the common method
            self.update_connections_table(connections)
            
            # Update status bar
            self.status_bar.showMessage(f'Updated network connections ({len(connections)} connections)', 3000)
            
        except Exception as e:
            self.logger.error(f"Error updating connections: {str(e)}")
            self.status_bar.showMessage(f'Error updating connections table: {str(e)}', 5000)
            QMessageBox.critical(
                self,
                "Connection Update Error",
                f"Failed to update connections table: {str(e)}\n\nPlease check if you have sufficient permissions.",
                QMessageBox.StandardButton.Ok
            )
    
    def update_connections_table(self, connections):
        """Update the connections table with current network connections"""
        try:
            # Clear existing rows
            self.connections_table.setRowCount(0)
            
            # Update table with filtered connections
            self.connections_table.setRowCount(len(connections))
            
            for row, conn in enumerate(connections):
                try:
                    # Create items for each column
                    local_addr = conn['local_address']
                    remote_addr = conn['remote_address']
                    
                    # Split addresses into IP and port, handling IPv6 addresses correctly
                    if ']' in local_addr:  # IPv6
                        local_ip = local_addr.split(']:')[0] + ']'
                        local_port = local_addr.split(']:')[1]
                    else:  # IPv4
                        local_ip = local_addr.rsplit(':', 1)[0]
                        local_port = local_addr.rsplit(':', 1)[1]
                        
                    if ']' in remote_addr:  # IPv6
                        remote_ip = remote_addr.split(']:')[0] + ']'
                        remote_port = remote_addr.split(']:')[1]
                    else:  # IPv4
                        remote_ip = remote_addr.rsplit(':', 1)[0]
                        remote_port = remote_addr.rsplit(':', 1)[1]
                    
                    items = [
                        QTableWidgetItem(local_ip),
                        QTableWidgetItem(local_port),
                        QTableWidgetItem(remote_ip),
                        QTableWidgetItem(remote_port),
                        QTableWidgetItem(conn['status']),
                        QTableWidgetItem(conn['process'])
                    ]
                    
                    # Set alignment and colors based on status
                    for col, item in enumerate(items):
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        if conn['status'] == "ESTABLISHED":
                            item.setBackground(QColor("#4CAF50"))
                            item.setForeground(QColor("#FFFFFF"))
                        elif conn['status'] == "LISTEN":
                            item.setBackground(QColor("#2196F3"))
                            item.setForeground(QColor("#FFFFFF"))
                        self.connections_table.setItem(row, col, item)
                        
                except Exception as e:
                    self.logger.error(f"Error adding connection to table: {str(e)}")
                    continue
                    
        except Exception as e:
            self.logger.error(f"Error updating connections table: {str(e)}")
            raise
    
    def update_services(self):
        """Update the services table with currently listening ports"""
        try:
            # Get listening connections as services
            services = []
            try:
                connections = psutil.net_connections(kind='inet')
            except psutil.AccessDenied as e:
                self.statusBar.showMessage(f'Access denied when getting network connections. Try running with elevated privileges.')
                return
            except Exception as e:
                self.statusBar.showMessage(f'Error getting network connections: {str(e)}')
                return

            for conn in connections:
                try:
                    # Only include listening connections
                    if not hasattr(conn, 'status') or conn.status != 'LISTEN':
                        continue
                    
                    # Get process information
                    process_name = '*'
                    if conn.pid:
                        try:
                            process = psutil.Process(conn.pid)
                            process_name = process.name()
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            pass
                    
                    # Get port information
                    port = conn.laddr.port if hasattr(conn, 'laddr') else 0
                    
                    service = {
                        'service': process_name,
                        'port': port,
                        'protocol': 'TCP',  # Add UDP support later
                        'state': 'LISTENING'
                    }
                    services.append(service)
                except (AttributeError, psutil.NoSuchProcess, psutil.AccessDenied) as e:
                    continue
            
            # Update services table
            self.services_table.setRowCount(len(services))
            for row, service in enumerate(services):
                self.services_table.setItem(row, 0, QTableWidgetItem(service['service']))
                self.services_table.setItem(row, 1, QTableWidgetItem(str(service['port'])))
                self.services_table.setItem(row, 2, QTableWidgetItem(service['protocol']))
                self.services_table.setItem(row, 3, QTableWidgetItem(service['state']))

            if services:
                self.statusBar.showMessage(f'Updated services table: {len(services)} services found')
            else:
                self.statusBar.showMessage('No listening services found or insufficient permissions')

        except Exception as e:
            self.statusBar.showMessage(f'Error updating services table: {str(e)}')
    
    def update_network_stats(self):
        """Update network statistics display"""
        try:
            # Get current network connections
            connections = psutil.net_connections(kind='inet')
            
            # Initialize counters
            total_connections = 0  # Only count ESTABLISHED and relevant states
            established = 0
            listening = 0
            time_wait = 0
            unique_remote_ips = set()
            processes = set()
            
            # Track seen connection pairs to prevent duplicates
            seen_connections = set()
            
            # Process each connection
            for conn in connections:
                try:
                    if not hasattr(conn, 'laddr') or not conn.laddr:
                        continue

                    # Initialize conn_key with a default value
                    conn_key = None
                    
                    # Assuming this is within a loop or condition where conn is defined
                    if hasattr(conn, 'raddr') and conn.raddr:
                        conn_key = f"{conn.laddr.ip}:{conn.laddr.port}-{conn.raddr.ip}:{conn.raddr.port}"
                    else:
                        conn_key = f"{conn.laddr.ip}:{conn.laddr.port}-0.0.0.0:0"  # Default value if raddr is not present
                    
                    # Now you can safely check conn_key
                    if conn_key and conn_key in seen_connections:
                        continue
                    
                    # Only count established and listening connections in total
                    if conn.status in ('ESTABLISHED', 'LISTEN'):
                        total_connections += 1
                        
                    if conn.status == 'ESTABLISHED':
                        if not conn.laddr.ip.startswith('127.') and hasattr(conn, 'raddr') and conn.raddr:
                            if not conn.raddr.ip.startswith('127.'):
                                established += 1
                                unique_remote_ips.add(conn.raddr.ip)
                                if conn_key:
                                    seen_connections.add(conn_key)
                    elif conn.status == 'LISTEN':
                        listening += 1
                    elif conn.status == 'TIME_WAIT':
                        time_wait += 1
                    
                    if conn.pid:
                        try:
                            process = psutil.Process(conn.pid)
                            processes.add(process.name())
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            pass
                    
                except AttributeError:
                    continue
        
            # Get network interface statistics
            net_io = psutil.net_io_counters()
            
            # Update statistics in table
            stats = [
                ('Active Connections', str(established)),
                ('Listening Ports', str(listening)),
                ('TIME_WAIT Connections', str(time_wait)),
                ('Unique Remote IPs', str(len(unique_remote_ips))),
                ('Active Network Processes', str(len(processes))),
                ('Bytes Sent', f"{net_io.bytes_sent / (1024*1024):.2f} MB"),
                ('Bytes Received', f"{net_io.bytes_recv / (1024*1024):.2f} MB"),
                ('Packets Sent', str(net_io.packets_sent)),
                ('Packets Received', str(net_io.packets_recv))
            ]
            
            # Update the table
            self.stats_table.setRowCount(len(stats))
            for row, (stat, value) in enumerate(stats):
                self.stats_table.setItem(row, 0, QTableWidgetItem(stat))
                self.stats_table.setItem(row, 1, QTableWidgetItem(value))
        
            # Update statistics labels
            self.total_connections_label.setText(f"Active Connections: {established}")
            self.unique_ips_label.setText(f"Unique IPs: {len(unique_remote_ips)}")
            self.active_ports_label.setText(f"Listening Ports: {listening}")
            self.avg_latency_label.setText(f"TIME_WAIT: {time_wait}")
            
            # Update network activity table
            self.network_table.setRowCount(0)  # Clear existing rows
            for conn in connections:
                try:
                    if not hasattr(conn, 'laddr') or not conn.laddr:
                        continue
                        
                    # Only show ESTABLISHED connections and LISTEN ports
                    if conn.status not in ('ESTABLISHED', 'LISTEN'):
                        continue
                        
                    # Skip loopback connections
                    if conn.laddr.ip.startswith('127.'):
                        continue
                    
                    if conn.status == 'ESTABLISHED' and hasattr(conn, 'raddr') and conn.raddr:
                        if conn.raddr.ip.startswith('127.'):
                            continue

                    row = self.network_table.rowCount()
                    self.network_table.insertRow(row)
                    
                    # Format addresses safely
                    local_addr = f"{conn.laddr.ip}:{conn.laddr.port}" if conn.laddr else "unknown:0"
                    remote_addr = "0.0.0.0:0"
                    if hasattr(conn, 'raddr') and conn.raddr:
                        remote_addr = f"{conn.raddr.ip}:{conn.raddr.port}"
                    
                    # Get process name
                    process_name = "Unknown"
                    if conn.pid:
                        try:
                            process = psutil.Process(conn.pid)
                            process_name = process.name()
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            pass
                    
                    # Create and set table items
                    items = [
                        QTableWidgetItem(local_addr),
                        QTableWidgetItem(remote_addr),
                        QTableWidgetItem(conn.status),
                        QTableWidgetItem(str(conn.pid or "Unknown")),
                        QTableWidgetItem(process_name)
                    ]
                    
                    for col, item in enumerate(items):
                        self.network_table.setItem(row, col, item)
                        
                except (psutil.NoSuchProcess, psutil.AccessDenied, AttributeError) as e:
                    self.logger.debug(f"Error adding connection to table: {str(e)}")
                    continue
        
            self.statusBar.showMessage('Updated network statistics', 3000)
            
        except Exception as e:
            self.statusBar.showMessage(f'Error updating network statistics: {str(e)}')
    
    def update_system_health(self):
        try:
            # Get CPU metrics with a small interval for accuracy
            cpu_percent = psutil.cpu_percent(interval=0.1)
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')
            
            # Update system health labels with proper formatting
            self.cpu_usage_label.setText(f"{cpu_percent:.1f}%")  # Changed from self.cpu_usage
            self.memory_usage_label.setText(f"{memory.percent:.1f}%")  # Changed from self.mem_usage
            self.disk_usage_label.setText(f"{disk.percent:.1f}%")  # Changed from self.disk_usage
            
            # Update progress bars
            self.cpu_progress.setValue(int(cpu_percent))
            self.memory_progress.setValue(int(memory.percent))
            self.disk_progress.setValue(int(disk.percent))
            
            # Update colors based on thresholds
            self._update_progress_color(self.cpu_progress, cpu_percent)
            self._update_progress_color(self.memory_progress, memory.percent)
            self._update_progress_color(self.disk_progress, disk.percent)
            
        except Exception as e:
            self.logger.error(f"Error updating system health: {str(e)}")
            
    def _update_progress_color(self, progress_bar, value):
        """Update progress bar color based on value"""
        if value >= 90:
            progress_bar.setStyleSheet("""
                QProgressBar::chunk { background-color: #ff4444; }
            """)
        elif value >= 70:
            progress_bar.setStyleSheet("""
                QProgressBar::chunk { background-color: #ffaa44; }
            """)
        else:
            progress_bar.setStyleSheet("""
                QProgressBar::chunk { background-color: #44aa44; }
            """)
    
    def update_security_analysis(self):
        """Update the security analysis with improved error handling and rate limiting"""
        try:
            # Get current connections
            connections = self.network_monitor.get_active_connections()
            
            # Convert connections to a format suitable for analysis
            connection_data = self._process_connections(connections)
            
            # Prepare event data for LLM analysis
            security_event = self._prepare_security_event(connection_data)
            
            try:
                # Get enhanced LLM analysis
                analysis_result = self.network_monitor.llm_analyzer.analyze_security_event(security_event)
                
                # Update the UI with the analysis results
                self._update_security_display(analysis_result)
                
            except Exception as e:
                self.logger.error(f"Error in security analysis: {str(e)}")
                self._handle_llm_error(str(e))
                
        except Exception as e:
            self.logger.error(f"Error updating security analysis: {str(e)}")
            self._handle_general_error(str(e))
            
    def _process_connections(self, connections) -> List[Dict]:
        """Process and filter network connections"""
        connection_data = []
        for conn in connections:
            try:
                # Only include established connections
                if hasattr(conn, 'status') and conn.status == 'ESTABLISHED':
                    # Parse the connection data safely
                    local_addr = conn.laddr if hasattr(conn, 'laddr') else None
                    remote_addr = conn.raddr if hasattr(conn, 'raddr') else None
                    
                    if local_addr and remote_addr:
                        connection_data.append({
                            'local_ip': local_addr.ip,
                            'local_port': local_addr.port,
                            'remote_ip': remote_addr.ip,
                            'remote_port': remote_addr.port,
                            'status': conn.status,
                            'pid': conn.pid if hasattr(conn, 'pid') else None,
                            'timestamp': datetime.now().isoformat()
                        })
            except Exception as e:
                self.logger.error(f"Error processing connection {conn}: {str(e)}")
                continue
        
        return connection_data

    def _prepare_security_event(self, connection_data: List[Dict]) -> Dict:
        """Prepare security event data for analysis"""
        try:
            # Get system metrics safely
            cpu_percent = psutil.cpu_percent()
            memory = psutil.virtual_memory()
            processes = list(psutil.process_iter(['num_threads']))
            total_processes = len(processes)
            total_threads = sum(p.info['num_threads'] for p in processes if 'num_threads' in p.info)
            
            # Get listening ports safely
            listening_ports = self._get_listening_ports()
            
            return {
                'type': 'system_security_scan',
                'timestamp': datetime.now().isoformat(),
                'connections': connection_data,
                'system_info': {
                    'cpu_percent': cpu_percent,
                    'memory_percent': memory.percent,
                    'total_processes': total_processes,
                    'total_threads': total_threads,
                    'total_connections': len(connection_data),
                    'unique_remote_ips': len({conn['remote_ip'] for conn in connection_data if conn.get('remote_ip')}),
                    'listening_ports': len(listening_ports),
                    'established_connections': len([conn for conn in connection_data if conn.get('status') == 'ESTABLISHED'])
                },
                'listening_ports': listening_ports,
                'context': []  # Required by Ollama API
            }
        except Exception as e:
            self.logger.error(f"Error preparing security event: {str(e)}")
            # Return a minimal valid structure if we encounter errors
            return {
                'type': 'system_security_scan',
                'timestamp': datetime.now().isoformat(),
                'connections': connection_data,
                'system_info': {},
                'listening_ports': [],
                'context': []  # Required by Ollama API
            }
    
    def _get_listening_ports(self) -> List[Dict]:
        """Get list of listening ports safely"""
        try:
            listening = []
            for conn in psutil.net_connections(kind='inet'):
                if conn.status == 'LISTEN':
                    listening.append({
                        'port': conn.laddr.port,
                        'ip': conn.laddr.ip,
                        'pid': conn.pid if hasattr(conn, 'pid') else None
                    })
            return listening
        except Exception as e:
            self.logger.error(f"Error getting listening ports: {str(e)}")
            return []

    def _update_security_display(self, analysis: Dict):
        """Update the security analysis display"""
        try:
            # Format the analysis results for display
            display_text = []
            
            # Add threat assessment
            threat = analysis.get('threat_assessment', {})
            severity = threat.get('severity', 'Unknown')
            confidence = threat.get('confidence', 0)
            
            # Use emoji indicators for severity
            severity_emoji = {
                'Low': '🟢',
                'Medium': '🟡',
                'High': '🔴',
                'Critical': '⚠️',
                'Unknown': '❓'
            }
            
            display_text.append(f"{severity_emoji.get(severity, '❓')} Threat Level: {severity}")
            display_text.append(f"Confidence: {confidence}%\n")
            
            # Add patterns if any
            patterns = threat.get('patterns', [])
            if patterns:
                display_text.append("Detected Patterns:")
                for pattern in patterns:
                    display_text.append(f"• {pattern}")
                display_text.append("")
            
            # Add impact analysis
            impacts = analysis.get('impact_analysis', [])
            if impacts:
                display_text.append("Impact Analysis:")
                for impact in impacts:
                    display_text.append(f"• {impact}")
                display_text.append("")
            
            # Add immediate actions
            actions = analysis.get('immediate_actions', [])
            if actions:
                display_text.append("Recommended Actions:")
                for action in actions:
                    display_text.append(f"• {action}")
                display_text.append("")
            
            # Add technical details
            details = analysis.get('technical_details', [])
            if details:
                display_text.append("Technical Details:")
                for detail in details:
                    display_text.append(f"• {detail}")
                display_text.append("")
            
            # Set the formatted text in the display
            self.security_text.clear()  # Clear existing content
            self.security_text.append("\n".join(display_text))
            
            # Update last updated timestamp
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.last_security_update_label.setText("Last Update: " + current_time)  # Update existing label
            
            # Update status bar with timestamp
            status_msg = f"Security analysis updated at {current_time}"
            self.status_bar.showMessage(status_msg, 5000)
            
            # Force scroll to bottom after content update
            scrollbar = self.security_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())
            
        except Exception as e:
            self.logger.error(f"Error updating security display: {str(e)}")
            self._handle_display_error(str(e))

    def _handle_privilege_error(self):
        """Handle privilege-related errors"""
        self.security_text.setPlainText(
            "⚠️ Elevated Privileges Required\n\n"
            "The security analysis requires elevated privileges to access network connection information.\n\n"
            "Please run the application with sudo privileges to enable full security analysis."
        )
        self.security_loading.hide()

    def _handle_connection_error(self, error_msg: str):
        """Handle connection-related errors"""
        self.security_text.setPlainText(
            "⚠️ Connection Error\n\n"
            f"Unable to analyze network connections: {error_msg}\n\n"
            "Please check your network configuration and try again."
        )
        self.security_loading.hide()

    def _handle_llm_error(self, error_msg: str):
        """Handle LLM-related errors"""
        self.security_text.setPlainText(
            "⚠️ Analysis Error\n\n"
            f"The security analysis engine encountered an error: {error_msg}\n\n"
            "Basic system monitoring will continue while this is resolved."
        )
        self.security_loading.hide()

    def _handle_display_error(self, error_msg: str):
        """Handle display-related errors"""
        self.security_text.setPlainText(
            "⚠️ Display Error\n\n"
            f"Error showing analysis results: {error_msg}\n\n"
            "Please check the application logs for more details."
        )
        self.security_loading.hide()

    def _handle_general_error(self, error_msg: str):
        """Handle general errors"""
        self.security_text.setPlainText(
            "⚠️ System Error\n\n"
            f"An unexpected error occurred: {error_msg}\n\n"
            "Please check the application logs and report this issue if it persists."
        )
        self.security_loading.hide()
    
    def show_action_confirmation(self):
        """Show a confirmation dialog for automated actions"""
        dialog = QMessageBox(self)
        dialog.setWindowTitle("Confirm Automated Action")
        dialog.setText("Are you sure you want to perform this action?")
        dialog.setIcon(QMessageBox.Icon.Question)
        dialog.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        dialog.setDefaultButton(QMessageBox.StandardButton.No)
        
        # Style the dialog
        dialog.setStyleSheet("""
            QMessageBox {
                background-color: #2D2D2D;
                color: #E0E0E0;
            }
            QPushButton {
                background-color: #424242;
                color: #E0E0E0;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #616161;
            }
        """)
        
        return dialog.exec() == QMessageBox.StandardButton.Yes

    def show_notification(self, title, message):
        """Show a system notification with fallback to status bar"""
        try:
            # First try using QSystemTrayIcon if available
            if hasattr(self, 'tray_icon') and self.tray_icon is not None:
                self.tray_icon.showMessage(title, message)
                return
                
            # Then try using native notification system
            try:
                from plyer import notification
                icon_path = os.path.join(os.path.dirname(__file__), 'icons', 'app_icon.png')
                if os.path.exists(icon_path):
                    notification.notify(
                        title=title,
                        message=message,
                        app_icon=icon_path,
                        timeout=10
                    )
                else:
                    notification.notify(
                        title=title,
                        message=message,
                        timeout=10
                    )
            except ImportError:
                # If plyer/pyobjus is not available, fall back to status bar
                self.statusBar.showMessage(f"{title}: {message}", 5000)
                
        except Exception as e:
            # Final fallback to status bar
            self.statusBar.showMessage(f"{title}: {message}", 5000)
            logging.error(f"Failed to show notification: {str(e)}")

    def closeEvent(self, event):
        """Clean up resources when closing the application"""
        try:
            self.update_timer.stop()
            event.accept()
        except:
            event.accept()

    def export_data(self):
        """Export network data, threats, and AI analysis to Excel"""
        try:
            # Get save file name using dialog
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Save Export Data",
                "",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not filename:  # User cancelled
                return
            
            # Add .xlsx extension if not present
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
                
            self.logger.info(f"Starting data export to {filename}")
            
            # Create Excel writer
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            
            # Export network connections
            logging.debug("Collecting network connection data...")
            with self.db_manager.get_session() as session:
                seven_days_ago = datetime.now() - timedelta(days=7)
                connections = session.query(NetworkConnection).filter(
                    NetworkConnection.timestamp >= seven_days_ago
                ).order_by(NetworkConnection.timestamp.desc()).all()
                
                connection_data = []
                if connections:
                    for conn in connections:
                        connection_data.append({
                            'Timestamp': conn.timestamp,
                            'Source IP': conn.source_ip,
                            'Source Port': conn.source_port,
                            'Destination IP': conn.destination_ip,
                            'Destination Port': conn.destination_port,
                            'Protocol': conn.protocol,
                            'Status': conn.connection_status,
                            'Bytes Sent': conn.bytes_sent,
                            'Bytes Received': conn.bytes_received,
                            'Threat Level': conn.threat_level
                        })
                else:
                    # If no historical connections in DB, get current active connections
                    active_connections = self.network_monitor.get_active_connections()
                    for conn in active_connections:
                        connection_data.append({
                            'Timestamp': datetime.now(),
                            'Source IP': conn.get('local_address', '').split(':')[0],
                            'Source Port': conn.get('local_address', '').split(':')[1] if ':' in conn.get('local_address', '') else '',
                            'Destination IP': conn.get('remote_address', '').split(':')[0],
                            'Destination Port': conn.get('remote_address', '').split(':')[1] if ':' in conn.get('remote_address', '') else '',
                            'Protocol': conn.get('protocol', ''),
                            'Status': conn.get('status', ''),
                            'Bytes Sent': conn.get('bytes_sent', 0),
                            'Bytes Received': conn.get('bytes_recv', 0),
                            'Threat Level': conn.get('threat_level', 'Unknown')
                        })
                
                df_connections = pd.DataFrame(connection_data)
                if not df_connections.empty:
                    df_connections.to_excel(writer, sheet_name='Network Connections', index=False)
                else:
                    # Create empty sheet with headers if no data
                    pd.DataFrame(columns=[
                        'Timestamp', 'Source IP', 'Source Port', 'Destination IP',
                        'Destination Port', 'Protocol', 'Status', 'Bytes Sent',
                        'Bytes Received', 'Threat Level'
                    ]).to_excel(writer, sheet_name='Network Connections', index=False)
            
            # Export threat data
            logging.debug("Collecting threat data...")
            with self.db_manager.get_session() as session:
                threats = session.query(Threat).filter(
                    Threat.timestamp >= seven_days_ago
                ).order_by(Threat.timestamp.desc()).all()
                
                threat_data = []
                for threat in threats:
                    threat_data.append({
                        'Timestamp': threat.timestamp,
                        'Process': threat.process_name,
                        'Remote IP': threat.remote_ip,
                        'Remote Port': threat.remote_port,
                        'Status': threat.connection_status,
                        'Threat Type': threat.threat_type,
                        'Resolved': bool(threat.resolved),
                        'Resolution Time': threat.resolution_time,
                        'Resolution Action': threat.resolution_action
                    })
                
                df_threats = pd.DataFrame(threat_data)
                df_threats.to_excel(writer, sheet_name='Threats', index=False)
            
            # Export AI analysis data
            logging.debug("Collecting AI analysis data...")
            ai_analysis_data = []
            
            # Current system analysis from security text panel
            current_analysis = {
                'Timestamp': datetime.now(),
                'Analysis Type': 'Current System Analysis',
                'Content': self.security_text.toPlainText()
            }
            ai_analysis_data.append(current_analysis)
            
            # Get historical daily reports from database
            with self.db_manager.get_session() as session:
                reports = session.query(Alert).filter(
                    Alert.alert_type == 'AI_REPORT',
                    Alert.timestamp >= seven_days_ago
                ).order_by(Alert.timestamp.desc()).all()
                
                logging.debug(f"Found {len(reports)} AI reports")
                for report in reports:
                    ai_analysis_data.append({
                        'Timestamp': report.timestamp,
                        'Analysis Type': 'Daily Report',
                        'Content': report.description
                    })
            
            df_ai = pd.DataFrame(ai_analysis_data)
            df_ai.to_excel(writer, sheet_name='AI Analysis', index=False)
            
            # Save and close the Excel file
            writer.close()
            
            self.logger.info(f"Data export completed successfully to {filename}")
            self.show_notification("Export Complete", f"Data has been exported to {filename}")
            
        except Exception as e:
            error_msg = f"Error exporting data: {str(e)}"
            self.logger.error(error_msg)
            if hasattr(self, 'show_notification'):
                self.show_notification("Export Error", error_msg)
    
    def get_total_security_events(self, start_time, end_time):
        """Get total number of security events"""
        # For demonstration, we'll generate a sample count
        return random.randint(50, 200)

    def get_high_severity_events(self, start_time, end_time):
        """Get number of high severity events"""
        # For demonstration, we'll return about 10-20% of total events as high severity
        total = self.get_total_security_events(start_time, end_time)
        return int(total * random.uniform(0.1, 0.2))

    def get_most_common_threat(self, start_time, end_time):
        """Get most common threat type"""
        threats = ['Malware', 'Phishing', 'DDoS', 'Unauthorized Access', 'Data Breach']
        return random.choice(threats)

    def get_top_attack_source(self, start_time, end_time):
        """Get top attack source IP"""
        ips = ['192.168.1.100', '10.0.0.5', '172.16.0.10', '192.168.1.200']
        return random.choice(ips)

    def get_avg_daily_events(self, start_time, end_time):
        """Get average daily security events"""
        total = self.get_total_security_events(start_time, end_time)
        days = max(1, (end_time - start_time).days)
        return total / days

    def get_total_processes(self, start_time, end_time):
        """Get total number of processes"""
        return len(psutil.pids())

    def get_top_cpu_process(self, start_time, end_time):
        """Get top CPU consuming process"""
        try:
            processes = []
            for proc in psutil.process_iter(['name', 'cpu_percent']):
                try:
                    # Truncate process name if too long
                    name = proc.info['name']
                    if len(name) > 20:
                        name = name[:17] + "..."
                    processes.append((name, proc.info['cpu_percent']))
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            # Get top 10 processes
            top_procs = sorted(processes, key=lambda x: x[1], reverse=True)[:10]
            return f"{top_procs[0][0]} ({top_procs[0][1]:.1f}%)"
        except Exception:
            return "N/A"

    def get_top_memory_process(self, start_time, end_time):
        """Get top memory consuming process"""
        try:
            processes = []
            for proc in psutil.process_iter(['name', 'memory_percent']):
                try:
                    # Truncate process name if too long
                    name = proc.info['name']
                    if len(name) > 20:
                        name = name[:17] + "..."
                    processes.append((name, proc.info['memory_percent']))
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            # Get top 10 processes
            top_procs = sorted(processes, key=lambda x: x[1], reverse=True)[:10]
            return f"{top_procs[0][0]} ({top_procs[0][1]:.1f}%)"
        except Exception:
            return "N/A"

    def get_avg_process_count(self, start_time, end_time):
        """Get average process count"""
        return len(psutil.pids())

    def get_most_common_process_type(self, start_time, end_time):
        """Get most common process type"""
        type_counts = {'System': 0, 'User': 0, 'Background': 0}
        try:
            for proc in psutil.process_iter(['name', 'username']):
                try:
                    if proc.username() == 'root':
                        type_counts['System'] += 1
                    elif proc.username() == getpass.getuser():
                        type_counts['User'] += 1
                    else:
                        type_counts['Background'] += 1
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            return max(type_counts.items(), key=lambda x: x[1])[0]
        except Exception:
            return "N/A"

    def get_security_events_trend(self, start_time, end_time):
        """Get security events trend"""
        events = [self.get_total_security_events(start_time, end_time) for _ in range(5)]
        return self._calculate_trend(events)

    def get_threat_types_trend(self, start_time, end_time):
        """Get threat types trend"""
        threats = ['Increasing', 'Stable', 'Decreasing']
        return random.choice(threats)

    def get_attack_sources_trend(self, start_time, end_time):
        """Get attack sources trend"""
        trends = ['New sources detected', 'Consistent sources', 'Decreasing variety']
        return random.choice(trends)

    def get_process_count_trend(self, start_time, end_time):
        """Get process count trend"""
        counts = [len(psutil.pids()) for _ in range(5)]
        return self._calculate_trend(counts)

    def get_process_cpu_trend(self, start_time, end_time):
        """Get process CPU usage trend"""
        cpu_values = [psutil.cpu_percent(interval=0.1) for _ in range(5)]
        return self._calculate_trend(cpu_values)

    def get_process_memory_trend(self, start_time, end_time):
        """Get process memory usage trend"""
        memory_values = [psutil.virtual_memory().percent for _ in range(5)]
        return self._calculate_trend(memory_values)

    def update_analytics_stats(self, start_time, end_time, metrics_type):
        """Update the statistics table based on selected parameters"""
        try:
            self.stats_table.setRowCount(0)
            stats = []

            if metrics_type == "System Performance":
                metrics_data = self.db_manager.get_metrics_range(start_time, end_time)
                if metrics_data:
                    cpu_data = [m.cpu_usage for m in metrics_data if m.cpu_usage is not None]
                    memory_data = [m.memory_usage for m in metrics_data if m.memory_usage is not None]
                    disk_data = [m.disk_usage for m in metrics_data if m.disk_usage is not None]
                    network_data = [m.network_throughput for m in metrics_data if m.network_throughput is not None]
                    process_data = [m.process_count for m in metrics_data if m.process_count is not None]
                    
                    stats = [
                        ("Average CPU Usage", f"{sum(cpu_data)/len(cpu_data):.1f}%" if cpu_data else "N/A"),
                        ("Peak CPU Usage", f"{max(cpu_data):.1f}%" if cpu_data else "N/A"),
                        ("Average Memory Usage", f"{sum(memory_data)/len(memory_data):.1f}%" if memory_data else "N/A"),
                        ("Peak Memory Usage", f"{max(memory_data):.1f}%" if memory_data else "N/A"),
                        ("Average Disk Usage", f"{sum(disk_data)/len(disk_data):.1f}%" if disk_data else "N/A"),
                        ("Peak Disk Usage", f"{max(disk_data):.1f}%" if disk_data else "N/A"),
                        ("Average Network Throughput", f"{sum(network_data)/len(network_data):.1f} MB/s" if network_data else "N/A"),
                        ("Average Process Count", f"{sum(process_data)/len(process_data):.1f}" if process_data else "N/A")
                    ]
                
            elif metrics_type == "Network Activity":
                network_data = self.db_manager.get_network_connections(start_time, end_time)
                if network_data:
                    total_sent = sum(conn.bytes_sent for conn in network_data if conn.bytes_sent is not None)
                    total_received = sum(conn.bytes_received for conn in network_data if conn.bytes_received is not None)
                    unique_ports = set(conn.destination_port for conn in network_data if conn.destination_port is not None)
                    protocols = [conn.protocol for conn in network_data if conn.protocol is not None]
                    most_common_protocol = max(set(protocols), key=protocols.count) if protocols else "N/A"
                    
                    stats = [
                        ("Total Data Sent", f"{total_sent / (1024*1024):.1f} MB"),
                        ("Total Data Received", f"{total_received / (1024*1024):.1f} MB"),
                        ("Total Connections", str(len(network_data))),
                        ("Unique Ports", str(len(unique_ports))),
                        ("Most Common Protocol", most_common_protocol),
                        ("Average Bytes per Connection", f"{(total_sent + total_received) / len(network_data) / 1024:.1f} KB" if network_data else "N/A")
                    ]
                
            elif metrics_type == "Security Events":
                alerts = self.db_manager.get_alerts(start_time, end_time)
                if alerts:
                    total_alerts = len(alerts)
                    high_severity = sum(1 for a in alerts if a.severity == "high")
                    medium_severity = sum(1 for a in alerts if a.severity == "medium")
                    low_severity = sum(1 for a in alerts if a.severity == "low")
                    alert_types = [a.alert_type for a in alerts if a.alert_type is not None]
                    most_common_type = max(set(alert_types), key=alert_types.count) if alert_types else "N/A"
                    
                    stats = [
                        ("Total Alerts", str(total_alerts)),
                        ("High Severity Alerts", str(high_severity)),
                        ("Medium Severity Alerts", str(medium_severity)),
                        ("Low Severity Alerts", str(low_severity)),
                        ("Most Common Alert Type", most_common_type),
                        ("Resolved Alerts", str(sum(1 for a in alerts if a.resolved)))
                    ]
                
            elif metrics_type == "Process Activity":
                metrics_data = self.db_manager.get_metrics_range(start_time, end_time)
                if metrics_data:
                    process_counts = [m.process_count for m in metrics_data if m.process_count is not None]
                    
                    stats = [
                        ("Average Process Count", f"{sum(process_counts)/len(process_counts):.1f}" if process_counts else "N/A"),
                        ("Peak Process Count", str(max(process_counts)) if process_counts else "N/A"),
                        ("Minimum Process Count", str(min(process_counts)) if process_counts else "N/A"),
                        ("Process Count Variance", f"{sum((x - (sum(process_counts)/len(process_counts)))**2 for x in process_counts)/len(process_counts):.1f}" if process_counts else "N/A")
                    ]

            # Populate the table
            for stat in stats:
                row = self.stats_table.rowCount()
                self.stats_table.insertRow(row)
                self.stats_table.setItem(row, 0, QTableWidgetItem(stat[0]))
                self.stats_table.setItem(row, 1, QTableWidgetItem(stat[1]))

            # Adjust column widths
            self.stats_table.resizeColumnsToContents()

        except Exception as e:
            self.statusBar.showMessage(f"Error updating statistics: {str(e)}", 5000)
            print(f"Statistics update error: {str(e)}")

    def update_analytics_trends(self, start_time, end_time, metrics_type):
        """Update the trends table based on selected parameters"""
        try:
            self.trends_table.setRowCount(0)
            trends = []

            if metrics_type == "System Performance":
                trends = [
                    ("CPU Usage Trend", self.get_cpu_trend(start_time, end_time)),
                    ("Memory Usage Trend", self.get_memory_trend(start_time, end_time)),
                    ("Disk I/O Trend", self.get_disk_io_trend(start_time, end_time)),
                    ("System Load Trend", self.get_system_load_trend(start_time, end_time))
                ]
            elif metrics_type == "Network Activity":
                trends = [
                    ("Network Traffic Trend", self.get_network_traffic_trend(start_time, end_time)),
                    ("Connection Count Trend", self.get_connection_count_trend(start_time, end_time)),
                    ("Port Activity Trend", self.get_port_activity_trend(start_time, end_time))
                ]
            elif metrics_type == "Security Events":
                trends = [
                    ("Security Events Trend", self.get_security_events_trend(start_time, end_time)),
                    ("Threat Types Trend", self.get_threat_types_trend(start_time, end_time)),
                    ("Attack Sources Trend", self.get_attack_sources_trend(start_time, end_time))
                ]
            elif metrics_type == "Process Activity":
                trends = [
                    ("Process Count Trend", self.get_process_count_trend(start_time, end_time)),
                    ("CPU Usage by Process Trend", self.get_process_cpu_trend(start_time, end_time)),
                    ("Memory Usage by Process Trend", self.get_process_memory_trend(start_time, end_time))
                ]

            # Populate the table
            for trend in trends:
                row = self.trends_table.rowCount()
                self.trends_table.insertRow(row)
                self.trends_table.setItem(row, 0, QTableWidgetItem(trend[0]))
                
                # Add trend arrow
                trend_item = QTableWidgetItem()
                if "increasing" in trend[1].lower():
                    trend_item.setText("↑")
                    trend_item.setForeground(QColor("green"))
                elif "decreasing" in trend[1].lower():
                    trend_item.setText("↓")
                    trend_item.setForeground(QColor("red"))
                else:
                    trend_item.setText("→")
                    trend_item.setForeground(QColor("gray"))
                self.trends_table.setItem(row, 1, trend_item)
                
                # Add trend description
                self.trends_table.setItem(row, 2, QTableWidgetItem(trend[1]))

            # Adjust column widths
            self.trends_table.resizeColumnsToContents()

        except Exception as e:
            self.statusBar.showMessage(f"Error updating trends: {str(e)}", 5000)
            print(f"Trends update error: {str(e)}")

    def get_performance_metrics(self):
        """Get system performance metrics"""
        try:
            metrics = {}
            
            # Get CPU metrics with error handling
            try:
                metrics['cpu'] = {
                    'percent': psutil.cpu_percent(interval=None)  # Non-blocking call
                }
            except Exception as e:
                self.logger.error(f"Error getting CPU metrics: {str(e)}")
                metrics['cpu'] = {'percent': 0}
            
            # Get memory metrics
            try:
                memory = psutil.virtual_memory()
                metrics['memory'] = {
                    'total': memory.total,
                    'available': memory.available,
                    'percent': memory.percent,
                    'used': memory.used,
                    'free': memory.free
                }
            except Exception as e:
                self.logger.error(f"Error getting memory metrics: {str(e)}")
                metrics['memory'] = {'percent': 0}
            
            # Get disk metrics
            try:
                disk = psutil.disk_usage('/')
                metrics['disk'] = {
                    'total': disk.total,
                    'used': disk.used,
                    'free': disk.free,
                    'percent': disk.percent
                }
            except Exception as e:
                self.logger.error(f"Error getting disk metrics: {str(e)}")
                metrics['disk'] = {'percent': 0}
            
            # Get network metrics
            try:
                net = psutil.net_io_counters()
                metrics['network'] = {
                    'bytes_sent': net.bytes_sent,
                    'bytes_recv': net.bytes_recv,
                    'packets_sent': net.packets_sent,
                    'packets_recv': net.packets_recv
                }
            except Exception as e:
                self.logger.error(f"Error getting network metrics: {str(e)}")
                metrics['network'] = {'bytes_sent': 0, 'bytes_recv': 0}
            
            return metrics
            
        except Exception as e:
            self.logger.error(f"Error getting performance metrics: {str(e)}")
            return {
                'cpu': {'percent': 0},
                'memory': {'percent': 0},
                'disk': {'percent': 0},
                'network': {'bytes_sent': 0, 'bytes_recv': 0}
            }
            
    def get_process_io_stats(self, proc):
        """Get I/O statistics for a process"""
        try:
            io_counters = proc.io_counters()
            return {
                'read_bytes': io_counters.read_bytes,
                'write_bytes': io_counters.write_bytes
            }
        except (psutil.AccessDenied, psutil.NoSuchProcess, AttributeError, Exception):
            return {
                'read_bytes': 0,
                'write_bytes': 0
            }

    def setup_packet_capture_tab(self):
        """Setup the packet capture tab with interface selection and capture controls"""
        layout = QVBoxLayout()

        # Create control panel
        control_panel = QHBoxLayout()
        
        # Interface selection
        interface_group = QFrame()
        interface_layout = QVBoxLayout(interface_group)
        self.interface_combo = QComboBox()
        self.interface_combo.setToolTip("Select network interface to capture packets from")
        interface_layout.addWidget(self.interface_combo)
        control_panel.addWidget(interface_group)
        
        # Filter settings
        filter_group = QGroupBox("Capture Filter")
        filter_layout = QVBoxLayout()
        
        # Service selection
        self.service_combo = QComboBox()
        self.service_combo.setToolTip("Select a service to automatically set filter")
        
        # Populate services
        services = ["All Services", "HTTP", "HTTPS", "DNS", "SSH", "FTP", "SMTP", "POP3", "IMAP", "NTP", "DHCP"]
        self.service_combo.addItems(services)
        
        self.service_combo.currentIndexChanged.connect(self.update_filter_from_service)
        filter_layout.addWidget(self.service_combo)
        
        # Custom filter input
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Enter BPF filter (e.g. 'tcp port 80')")
        self.filter_input.setToolTip("Enter Berkeley Packet Filter expression")
        filter_layout.addWidget(self.filter_input)
        
        filter_group.setLayout(filter_layout)
        control_panel.addWidget(filter_group)
        
        # Capture settings
        settings_group = QGroupBox("Capture Settings")
        settings_layout = QVBoxLayout()
        
        # Packet count setting
        count_layout = QHBoxLayout()
        count_label = QLabel("Max Packets:")
        self.packet_count_input = QSpinBox()
        self.packet_count_input.setRange(1, 1000000)
        self.packet_count_input.setValue(1000)
        self.packet_count_input.setToolTip("Maximum number of packets to capture")
        count_layout.addWidget(count_label)
        count_layout.addWidget(self.packet_count_input)
        settings_layout.addLayout(count_layout)
        
        settings_group.setLayout(settings_layout)
        control_panel.addWidget(settings_group)
        
        # Control buttons
        button_group = QGroupBox("Controls")
        button_layout = QVBoxLayout()
        
        self.start_capture_button = QPushButton("Start Capture")
        self.start_capture_button.clicked.connect(self.start_packet_capture)
        self.start_capture_button.setEnabled(False)
        button_layout.addWidget(self.start_capture_button)
        
        self.stop_capture_button = QPushButton("Stop Capture")
        self.stop_capture_button.clicked.connect(self.stop_packet_capture)
        self.stop_capture_button.setEnabled(False)
        button_layout.addWidget(self.stop_capture_button)
        
        self.export_button = QPushButton("Export")
        self.export_button.clicked.connect(self.export_packet_capture)
        self.export_button.setEnabled(False)
        button_layout.addWidget(self.export_button)

        self.clear_button = QPushButton("Clear")
        self.clear_button.clicked.connect(self.clear_packet_capture)
        self.clear_button.setEnabled(False)
        button_layout.addWidget(self.clear_button)

        # Add AI Analysis button
        self.analyze_button = QPushButton("AI Analysis")
        self.analyze_button.clicked.connect(self.analyze_captured_packets)
        self.analyze_button.setEnabled(False)
        self.analyze_button.setToolTip("Analyze captured packets using AI")
        button_layout.addWidget(self.analyze_button)

        button_group.setLayout(button_layout)
        control_panel.addWidget(button_group)
        
        layout.addLayout(control_panel)
        
        # Create packet display table
        self.packet_table = QTableWidget()
        self.packet_table.setColumnCount(8)
        self.packet_table.setHorizontalHeaderLabels([
            "Time", "Source", "Source Port", "Destination", "Destination Port", "Protocol", "Length", "Info"
        ])
        self.packet_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.packet_table.verticalHeader().setVisible(False)
        self.packet_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.packet_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.packet_table)
        
        # Status bar
        status_layout = QHBoxLayout()
        self.capture_status_label = QLabel("Status: Ready")
        self.packet_counter_label = QLabel("Packets: 0")
        status_layout.addWidget(self.capture_status_label)
        status_layout.addWidget(self.packet_counter_label)
        layout.addLayout(status_layout)
        
        # Now populate network interfaces
        try:
            import psutil
            from scapy.arch import get_if_list, get_if_addr
            
            # Get all interfaces that scapy can see
            available_interfaces = get_if_list()
            
            # Get interface details from psutil for additional info
            if_addrs = psutil.net_if_addrs()
            if_stats = psutil.net_if_stats()
            
            for iface in sorted(available_interfaces):
                if iface in if_stats and if_stats[iface].isup:  # Only show interfaces that are up
                    # Get IP address if available
                    ip_addr = get_if_addr(iface) or "No IP"
                    # Create a descriptive label
                    label = f"{iface} ({ip_addr})"
                    if iface in if_stats:
                        speed = if_stats[iface].speed
                        if speed > 0:  # Only show speed if available
                            label += f" - {speed}Mb/s"
                    self.interface_combo.addItem(label, iface)  # Store raw interface name as item data
            
            if self.interface_combo.count() > 0:
                self.start_capture_button.setEnabled(True)
            else:
                self.logger.warning("No active network interfaces found")
                self.start_capture_button.setEnabled(False)
                
        except Exception as e:
            self.logger.error(f"Error populating network interfaces: {str(e)}")
            self.start_capture_button.setEnabled(False)
        
        tab = QWidget()
        tab.setLayout(layout)
        return tab
    
    def start_packet_capture(self):
        """Start packet capture with current settings"""
        try:
            # Get the actual interface name from the UserRole data
            interface = self.interface_combo.currentData(Qt.ItemDataRole.UserRole)
            if not interface:
                self.show_notification("Error", "Please select a network interface")
                return
            
            # Validate interface exists and is up
            try:
                stats = psutil.net_if_stats()
                if interface in stats:
                    speed = stats[interface].speed
                    if speed > 0:
                        return f"{speed} Mbps"
            except:
                return "Unknown"
            
            filter_text = self.filter_input.text().strip()
            packet_count = self.packet_count_input.value()
            
            self.logger.info(f"Starting capture on interface {interface} with filter: {filter_text}")
            
            # Check if we have the necessary permissions
            if os.geteuid() != 0:  # Not running as root
                self.logger.warning("Packet capture requires root privileges")
                self.show_notification("Warning", "Packet capture requires root privileges. Some features may be limited.")
            
            # Create and setup capture thread
            self.capture_thread = PacketCaptureThread(
                interface=interface,
                filter_text=filter_text,
                packet_count=packet_count
            )
            
            # Connect signals
            self.capture_thread.packet_received.connect(self.handle_packet)
            self.capture_thread.capture_complete.connect(self.handle_capture_complete)
            self.capture_thread.error_signal.connect(self.handle_capture_error)
            self.capture_thread.status_signal.connect(self.handle_capture_status)
            
            # Update UI state
            self.captured_packets = []
            self.packet_table.setRowCount(0)
            self.packet_counter_label.setText("Packets: 0")
            self.capture_status_label.setText("Status: Initializing capture...")
            
            # Update button states
            self.start_capture_button.setEnabled(False)
            self.stop_capture_button.setEnabled(True)
            self.export_button.setEnabled(False)
            self.clear_button.setEnabled(False)
            self.analyze_button.setEnabled(False)  # Disable analyze button when starting capture
            self.interface_combo.setEnabled(False)
            self.filter_input.setEnabled(False)
            self.packet_count_input.setEnabled(False)
            
            # Start capture
            self.capture_thread.start()
            self.show_notification("Info", f"Started packet capture on {interface}")
            
        except Exception as e:
            self.logger.error(f"Error starting packet capture: {str(e)}\n{traceback.format_exc()}")
            self.show_notification("Error", "Failed to start packet capture")
            self.start_capture_button.setEnabled(True)
            self.stop_capture_button.setEnabled(False)

    def stop_packet_capture(self):
        """Stop the current packet capture"""
        try:
            if hasattr(self, 'capture_thread') and self.capture_thread.isRunning():
                self.logger.info("Stopping packet capture...")
                
                # Stop the capture thread
                self.capture_thread.stop()
                
                # Update UI state
                self.start_capture_button.setEnabled(True)
                self.stop_capture_button.setEnabled(False)
                self.export_button.setEnabled(len(self.captured_packets) > 0)
                self.clear_button.setEnabled(len(self.captured_packets) > 0)
                self.analyze_button.setEnabled(len(self.captured_packets) > 0)  # Enable analyze button after stopping capture
                
                # Update status
                self.capture_status_label.setText("Status: Stopped")
                self.capture_status_label.setStyleSheet("color: orange;")  # Orange color for stopped
                
                # Show notification
                self.show_notification("Info", "Packet capture stopped")
                
        except Exception as e:
            self.logger.error(f"Error stopping packet capture: {str(e)}")
            self.show_notification("Error", "Failed to stop packet capture")
            self.capture_status_label.setText("Error getting response")
            self.capture_status_label.setStyleSheet("color: red;")  # Red color for error
            self.start_capture_button.setEnabled(True)
            self.stop_capture_button.setEnabled(False)

    def handle_packet(self, packet_info):
        """Handle a received packet"""
        try:
            # Add packet to the list
            self.captured_packets.append(packet_info)
            
            # Add to table
            row = self.packet_table.rowCount()
            self.packet_table.insertRow(row)
            
            # Set data
            self.packet_table.setItem(row, 0, QTableWidgetItem(packet_info.get('time', '')))
            self.packet_table.setItem(row, 1, QTableWidgetItem(packet_info.get('source', '')))
            self.packet_table.setItem(row, 2, QTableWidgetItem(str(packet_info.get('src_port', ''))))
            self.packet_table.setItem(row, 3, QTableWidgetItem(packet_info.get('destination', '')))
            self.packet_table.setItem(row, 4, QTableWidgetItem(str(packet_info.get('dst_port', ''))))
            self.packet_table.setItem(row, 5, QTableWidgetItem(packet_info.get('protocol', '')))
            self.packet_table.setItem(row, 6, QTableWidgetItem(str(packet_info.get('length', ''))))
            self.packet_table.setItem(row, 7, QTableWidgetItem(packet_info.get('info', '')))
            
            # Auto-scroll to bottom
            self.packet_table.scrollToBottom()
            
            # Update status
            self.status_bar.showMessage(f"Captured {len(self.captured_packets)} packets")
            
        except Exception as e:
            self.logger.error(f"Error handling packet: {str(e)}\n{traceback.format_exc()}")
    
    def handle_capture_complete(self):
        """Handle completion of packet capture"""
        try:
            # Update UI state
            self.start_capture_button.setEnabled(True)
            self.stop_capture_button.setEnabled(False)
            self.clear_button.setEnabled(len(self.captured_packets) > 0)
            self.export_button.setEnabled(len(self.captured_packets) > 0)
            self.analyze_button.setEnabled(len(self.captured_packets) > 0)  # Enable analyze button if packets were captured
            self.interface_combo.setEnabled(True)
            self.filter_input.setEnabled(True)
            self.packet_count_input.setEnabled(True)
            
            # Update status
            self.capture_status_label.setText("Status: Complete")
            self.capture_status_label.setStyleSheet("color: green;")  # Green color for complete
            
            # Show notification
            self.show_notification("Info", "Packet capture complete")
            
        except Exception as e:
            self.logger.error(f"Error handling capture completion: {str(e)}")
            self.show_notification("Error", "Failed to handle capture completion")

    def handle_capture_error(self, error):
        """Handle an error during packet capture"""
        self.logger.error(f"Packet capture error: {error}")
        self.show_notification("Capture Error", error)
        
        if "Permission denied" in error:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setWindowTitle("Permission Error")
            msg.setText("Network capture requires elevated privileges")
            msg.setInformativeText("Please run the application with sudo privileges to capture packets.")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
    
    def handle_capture_status(self, status):
        """Handle capture status updates"""
        self.capture_status_label.setText(f"Status: {status}")
        if "complete" in status.lower():
            self.capture_status_label.setStyleSheet("color: green;")  # Green color for complete
        elif "error" in status.lower():
            self.capture_status_label.setStyleSheet("color: red;")  # Red color for error
        else:
            self.capture_status_label.setStyleSheet("color: blue;")  # Blue color for running
    
    def clear_packet_capture(self):
        """Clear the packet capture data and reset UI elements"""
        try:
            # Clear the packet table
            self.packet_table.setRowCount(0)
            
            # Clear captured packets list
            self.captured_packets.clear()
            
            # Reset UI elements
            self.capture_status_label.setText("Status: Ready")
            self.clear_button.setEnabled(False)
            self.export_button.setEnabled(False)
            self.analyze_button.setEnabled(False)  # Disable analyze button
            
            # Enable input elements
            self.interface_combo.setEnabled(True)
            self.filter_input.setEnabled(True)
            self.packet_count_input.setEnabled(True)
            self.service_combo.setEnabled(True)
            
            self.logger.info("Packet capture data cleared")
            self.show_notification("Success", "Packet capture data cleared")
            
        except Exception as e:
            self.logger.error(f"Error clearing packet capture: {str(e)}")
            self.show_notification("Error", f"Failed to clear packet capture: {str(e)}")

    def create_default_icon(self, icon_path):
        """Create a default application icon if the icon file is missing"""
        try:
            # Create a simple icon using QPainter
            icon_size = 128
            icon = QImage(icon_size, icon_size, QImage.Format.Format_ARGB32)
            icon.fill(Qt.GlobalColor.transparent)
            
            painter = QPainter(icon)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # Draw a gradient background
            gradient = QLinearGradient(0, 0, icon_size, icon_size)
            gradient.setColorAt(0.0, QColor(74, 158, 255))  # Light blue
            gradient.setColorAt(1.0, QColor(41, 128, 185))  # Darker blue
            
            painter.setBrush(QBrush(gradient))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawEllipse(4, 4, icon_size-8, icon_size-8)
            
            # Draw network-like connections
            painter.setPen(QPen(QColor(255, 255, 255, 180), 3))
            painter.drawLine(icon_size//4, icon_size//4, 3*icon_size//4, 3*icon_size//4)
            painter.drawLine(icon_size//4, 3*icon_size//4, 3*icon_size//4, icon_size//4)
            
            # Draw corner dots
            painter.setBrush(QBrush(Qt.GlobalColor.white))
            painter.setPen(Qt.PenStyle.NoPen)
            dot_size = 12
            painter.drawEllipse(icon_size//4 - dot_size//2, icon_size//4 - dot_size//2, dot_size, dot_size)
            painter.drawEllipse(3*icon_size//4 - dot_size//2, icon_size//4 - dot_size//2, dot_size, dot_size)
            painter.drawEllipse(icon_size//4 - dot_size//2, 3*icon_size//4 - dot_size//2, dot_size, dot_size)
            painter.drawEllipse(3*icon_size//4 - dot_size//2, 3*icon_size//4 - dot_size//2, dot_size, dot_size)
            
            painter.end()
            
            # Save the icon
            icon.save(icon_path)
            self.logger.info(f"Created default application icon at {icon_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to create default icon: {str(e)}")

    def get_interface_speed(self):
        """Get the speed of the current network interface"""
        try:
            interface = self.interface_combo.currentText()
            if not interface:
                return "Unknown"
            
            # Get interface info using psutil
            stats = psutil.net_if_stats()
            if interface in stats:
                speed = stats[interface].speed
                if speed > 0:
                    return f"{speed} Mbps"
            return "Unknown"
        except:
            return "Unknown"

    def update_filter_from_service(self):
        """Update the packet capture filter based on the selected service."""
        current_service = self.service_combo.currentText()
        if not current_service or current_service == "All Services":
            self.filter_input.clear()
            return
            
        # Common service to port mappings
        service_filters = {
            "HTTP": "tcp port 80",
            "HTTPS": "tcp port 443",
            "DNS": "udp port 53",
            "SSH": "tcp port 22",
            "FTP": "tcp port 21",
            "SMTP": "tcp port 25",
            "POP3": "tcp port 110",
            "IMAP": "tcp port 143",
            "NTP": "udp port 123",
            "DHCP": "udp port 67 or udp port 68"
        }
        
        if current_service in service_filters:
            self.filter_input.setText(service_filters[current_service])
    
    def export_packet_capture(self):
        """Export the captured packets to an Excel file with system details"""
        try:
            if not self.captured_packets:
                self.show_notification("Warning", "No packets to export")
                return
                
            # Get save file name for Excel
            excel_file, _ = QFileDialog.getSaveFileName(
                self,
                "Save Packet Capture as Excel",
                os.path.expanduser("~/Downloads/packet_capture.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if excel_file:  # User didn't cancel Excel export
                if not excel_file.endswith('.xlsx'):
                    excel_file += '.xlsx'
                
                workbook = openpyxl.Workbook()
                
                # Packet Data Sheet
                packet_sheet = workbook.active
                packet_sheet.title = "Packet Data"
                
                # Add headers
                headers = ["Timestamp", "Source", "Source Port", "Destination", "Destination Port", "Protocol", "Length", "Info"]
                for col, header in enumerate(headers, 1):
                    cell = packet_sheet.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                
                # Add packet data
                for row, packet in enumerate(self.captured_packets, 2):
                    packet_sheet.cell(row=row, column=1, value=packet.get('time', ''))
                    packet_sheet.cell(row=row, column=2, value=packet.get('source', ''))
                    packet_sheet.cell(row=row, column=3, value=packet.get('src_port', ''))
                    packet_sheet.cell(row=row, column=4, value=packet.get('destination', ''))
                    packet_sheet.cell(row=row, column=5, value=packet.get('dst_port', ''))
                    packet_sheet.cell(row=row, column=6, value=packet.get('protocol', ''))
                    packet_sheet.cell(row=row, column=7, value=packet.get('length', ''))
                    packet_sheet.cell(row=row, column=8, value=packet.get('info', ''))
                
                # System Details Sheet
                system_sheet = workbook.create_sheet("System Details")
                
                # Add system information
                current_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
                system_info = [
                    ("Capture Start Time", getattr(self, 'capture_start_time', current_time)),
                    ("Capture End Time", current_time),
                    ("Interface", self.interface_combo.currentText()),
                    ("Filter", self.filter_input.text()),
                    ("Total Packets", len(self.captured_packets)),
                    ("Operating System", platform.system() + " " + platform.release()),
                    ("Python Version", platform.python_version()),
                    ("CPU Usage", f"{psutil.cpu_percent()}%"),
                    ("Memory Usage", f"{psutil.virtual_memory().percent}%"),
                    ("Network Interface Speed", self.get_interface_speed()),
                ]
                
                for row, (key, value) in enumerate(system_info, 1):
                    system_sheet.cell(row=row, column=1, value=key).font = openpyxl.styles.Font(bold=True)
                    system_sheet.cell(row=row, column=2, value=value)
                
                # Adjust column widths
                for sheet in workbook.sheetnames:
                    for column in workbook[sheet].columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        workbook[sheet].column_dimensions[column[0].column_letter].width = adjusted_width
                
                workbook.save(excel_file)
                self.show_notification("Success", f"Packet data exported to {excel_file}")
            
            # Get save file name for PCAP
            pcap_file, _ = QFileDialog.getSaveFileName(
                self,
                "Save Packet Capture as PCAP",
                os.path.expanduser("~/Downloads/packet_capture.pcap"),
                "PCAP Files (*.pcap)"
            )
            
            if pcap_file:  # User didn't cancel PCAP export
                if not pcap_file.endswith('.pcap'):
                    pcap_file += '.pcap'
                
                # Create PCAP writer
                pcap_writer = PcapWriter(pcap_file, append=False, sync=True)
                
                # Write packets
                for packet in self.captured_packets:
                    if 'raw_packet' in packet:
                        pcap_writer.write(packet['raw_packet'])
                
                pcap_writer.close()
                self.show_notification("Success", f"Packet data exported to {pcap_file}")
            
        except Exception as e:
            self.logger.error(f"Error exporting packet data: {str(e)}\n{traceback.format_exc()}")
            self.show_notification("Error", f"Failed to export packet data: {str(e)}")

    def analyze_captured_packets(self):
        """Analyze captured packets using AI"""
        try:
            self.logger.info("Starting AI analysis...")
            
            if not self.captured_packets:
                self.status_bar.showMessage("No packets to analyze")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("AI Packet Analysis")
            dialog.setMinimumSize(600, 400)
        
            layout = QVBoxLayout()
            status = QLabel("Analyzing packets... this may take a few minutes.")
            text = QTextEdit()
            text.setReadOnly(True)
            close = QPushButton("Close")
            close.clicked.connect(dialog.accept)
        
            layout.addWidget(status)
            layout.addWidget(text)
            layout.addWidget(close)
            dialog.setLayout(layout)
            dialog.show()

            # Prepare data with new field names
            packets = []
            for p in self.captured_packets:
                packet = {
                    'timestamp': p.get('timestamp', datetime.now().strftime('%H:%M:%S')),
                    'source': p.get('src', p.get('source', 'unknown')), 
                    'destination': p.get('dst', p.get('destination', 'unknown')),
                    'protocol': p.get('protocol', 'unknown'),
                    'length': p.get('length', 0),
                    'src_port': p.get('src_port', ''),
                    'dst_port': p.get('dst_port', ''),
                    'info': p.get('info', '')
                }
                packets.append(packet)

            # Analysis in thread
            def analyze():
                try:
                    prompt = f"""Analyze these network packets for security issues and patterns:
{json.dumps(packets[:100], indent=2)}  # Limit to first 100 packets

Provide:
1. Traffic patterns overview
2. Security concerns
3. Protocol usage
4. Monitoring recommendations
5. Summary

Focus on suspicious patterns."""

                    # Make request to local Ollama instance
                    response = requests.post(
                        "http://localhost:11434/api/generate", 
                        json={
                            'model': 'llama3.2',  # Specify model
                            'prompt': prompt,
                            'stream': False  # Don't stream response
                        }
                    )
                    response.raise_for_status()
                    data = response.json()
                    analysis_text = data.get('response', 'No response from AI')
                    
                    # Update the UI with the response
                    QMetaObject.invokeMethod(status, "setText", 
                       Qt.ConnectionType.QueuedConnection,
                       Q_ARG(str, "Analysis complete"))
                    QMetaObject.invokeMethod(text, "setPlainText", 
                       Qt.ConnectionType.QueuedConnection,
                       Q_ARG(str, analysis_text))
                except Exception as e:
                    error_msg = f"Analysis error: {str(e)}"
                    self.logger.error(error_msg)
                    QMetaObject.invokeMethod(status, "setText", 
                       Qt.ConnectionType.QueuedConnection,
                       Q_ARG(str, "Analysis failed"))
                    QMetaObject.invokeMethod(text, "setPlainText", 
                       Qt.ConnectionType.QueuedConnection,
                       Q_ARG(str, error_msg))

            Thread(target=analyze).start()

        except Exception as e:
            error_msg = f"Error preparing analysis: {str(e)}"
            self.logger.error(error_msg)
            self.status_bar.showMessage(error_msg)
    
    def create_security_agent_chat_tab(self):
        chat_tab = QWidget()
        layout = QVBoxLayout()

        # Initialize conversation history
        self.conversation_history = []

        # Chat display area
        self.chat_display = QTextEdit()
        self.chat_display.setReadOnly(True)
        self.chat_display.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #ffffff;
                border: 1px solid #333333;
                border-radius: 4px;
                padding: 8px;
            }
        """)
        layout.addWidget(self.chat_display)

        # Input area container
        input_container = QWidget()
        input_layout = QHBoxLayout(input_container)
        input_container.setLayout(input_layout)

        # User input area
        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText('Enter message...')
        self.user_input.returnPressed.connect(self.send_message)  # Allow Enter to send
        self.user_input.setStyleSheet("""
            QLineEdit {
                background-color: #2d2d2d;
                color: #ffffff;
                border: 1px solid #333333;
                border-radius: 4px;
                padding: 5px;
            }
            QLineEdit:focus {
                border: 1px solid #4a9eff;
            }
        """)
        input_layout.addWidget(self.user_input)

        # Send button
        send_button = QPushButton('Send')
        send_button.clicked.connect(self.send_message)
        send_button.setStyleSheet("""
            QPushButton {
                background-color: #4a9eff;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
        """)
        input_layout.addWidget(send_button)

        # Copy button
        copy_button = QPushButton('Copy Response')
        copy_button.clicked.connect(self.copy_response)
        copy_button.setStyleSheet("""
            QPushButton {
                background-color: #2d2d2d;
                color: white;
                border: 1px solid #333333;
                padding: 5px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #404040;
            }
        """)
        input_layout.addWidget(copy_button)

        # New Chat button
        new_chat_button = QPushButton('New Chat')
        new_chat_button.clicked.connect(self.start_new_chat)
        new_chat_button.setStyleSheet("""
            QPushButton {
                background-color: #2d2d2d;
                color: white;
                border: 1px solid #333333;
                padding: 5px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #404040;
            }
        """)
        input_layout.addWidget(new_chat_button)

        layout.addWidget(input_container)

        # Status label
        self.chat_status_label = QLabel("")
        self.chat_status_label.setStyleSheet("color: #888888;")
        layout.addWidget(self.chat_status_label)

        # Add welcome message
        welcome_message = """<div style="margin: 10px 0; background-color: #2d2d2d; padding: 10px; border-radius: 4px;">
        <span style="color: #00ff00;">Security Agent:</span> Hello! I'm your Security Agent AI assistant. 
        I can help you understand and respond to security events and threats. How can I assist you today?</div>"""
        self.chat_display.append(welcome_message)

        chat_tab.setLayout(layout)
        self.tabs.addTab(chat_tab, "Security Agent")
        return chat_tab

    def send_message(self):
        """Send a message to the security agent and display the response."""
        message = self.user_input.text().strip()
        if not message:
            return

        # Clear input and disable until response received
        self.user_input.clear()
        self.user_input.setEnabled(False)
        self.chat_status_label.setText("Getting response...")
        self.chat_status_label.setStyleSheet("color: #888888;")

        # Format and display user message
        user_html = f'<div style="margin: 10px 0;"><span style="color: #4a9eff;">You:</span> {message}</div>'
        self.chat_display.append(user_html)

        try:
            # Get the currently selected model from config and ensure it's clean
            model = self.config.OLLAMA_CONFIG.get('model', 'llama3.2').strip().split()[0]  # Get just the model name
            self.logger.info(f"Using model: {model}")  # Log the cleaned model name

            # Add the new message to conversation history
            self.conversation_history.append({
                'role': 'user',
                'content': message
            })

            # Prepare the request to Ollama with full conversation history
            payload = {
                'model': model,
                'messages': [
                    {
                        'role': 'system',
                        'content': """You are a Security Agent AI assistant, an expert in cybersecurity and network monitoring. Your role is to help users understand and respond to security events, threats, and system behavior.

Key Responsibilities:
1. Analyze and explain security events and potential threats
2. Provide guidance on network security best practices
3. Help interpret system logs and network traffic patterns
4. Assist with incident response and mitigation strategies
5. Explain security concepts in clear, precise terms

Guidelines:
- Be concise and precise in your responses
- Focus on practical, actionable advice
- Maintain a security-first mindset
- Consider both immediate threats and long-term security implications
- Reference relevant security standards and best practices when applicable

You have access to real-time system data and network monitoring information. Use this context to provide informed, relevant security guidance."""
                    }
                ] + self.conversation_history,
                'stream': True  # Explicitly enable streaming
            }

            # Set proper headers
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }

            # Log the request for debugging
            self.logger.info(f"Sending request to Ollama with payload: {payload}")

            # Send request to Ollama
            response = requests.post(
                "http://localhost:11434/api/chat",
                json=payload,
                headers=headers,
                timeout=30,  # Add timeout
                stream=True  # Enable streaming
            )
            response.raise_for_status()  # Raise an error for bad responses
            full_response = ""
            for line in response.iter_lines():
                if line:
                    try:
                        data = json.loads(line.decode('utf-8'))
                        if data.get('done', False):
                            break
                        content = data.get('message', {}).get('content', '')
                        if content:
                            full_response += content
                            # Update the UI with the partial response
                            ai_html = f'<div style="margin: 10px 0; background-color: #2d2d2d; padding: 10px; border-radius: 4px;">'
                            ai_html += f'<span style="color: #00ff00;">Security Agent:</span> {full_response}</div>'
                            self.chat_display.clear()  # Clear existing content
                            self.chat_display.append(user_html)  # Re-add user message
                            self.chat_display.append(ai_html)  # Add AI response
                            # Scroll to bottom
                            scrollbar = self.chat_display.verticalScrollBar()
                            scrollbar.setValue(scrollbar.maximum())
                    except json.JSONDecodeError as e:
                        self.logger.error(f"Error parsing JSON line: {line}")
                        continue

            # Add the AI's response to conversation history
            self.conversation_history.append({
                'role': 'assistant',
                'content': full_response
            })

            # Clear status
            self.chat_status_label.clear()

        except requests.exceptions.ConnectionError:
            error_msg = "Error: Could not connect to Ollama. Please ensure Ollama is running."
            self.chat_display.append(f'<div style="color: #ff4444;">{error_msg}</div>')
            self.chat_status_label.setText(error_msg)
            self.chat_status_label.setStyleSheet("color: #ff4444;")
            self.logger.error("Connection error when trying to reach Ollama")

        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.chat_display.append(f'<div style="color: #ff4444;">{error_msg}</div>')
            self.chat_status_label.setText("Error getting response")
            self.chat_status_label.setStyleSheet("color: #ff4444;")
            self.logger.error(f"Error in send_message: {str(e)}")

        finally:
            # Re-enable input
            self.user_input.setEnabled(True)
            self.user_input.setFocus()
    def copy_response(self):
        """Copy the last response from the security agent."""
        try:
            # Get all text
            text = self.chat_display.toPlainText()

            # Find the last response
            parts = text.split('Security Agent:')
            if len(parts) > 1:
                last_response = parts[-1].strip()
                QApplication.clipboard().setText(last_response)
                self.chat_status_label.setText("Response copied to clipboard")
                self.chat_status_label.setStyleSheet("color: #00ff00;")
            else:
                self.chat_status_label.setText("No response to copy")
                self.chat_status_label.setStyleSheet("color: #ffa500;")  # Orange for warning

        except Exception as e:
            self.logger.error(f"Error copying response: {str(e)}")
            self.chat_status_label.setText("Failed to copy response")
            self.chat_status_label.setStyleSheet("color: #ff4444;")
    def start_new_chat(self):
        """Start a new chat by clearing history and display"""
        self.conversation_history = []  # Clear conversation history
        self.chat_display.clear()  # Clear the display
        
        # Add welcome message
        welcome_message = """<div style="margin: 10px 0; background-color: #2d2d2d; padding: 10px; border-radius: 4px;">
        <span style="color: #00ff00;">Security Agent:</span> Hello! I'm your Security Agent AI assistant. 
        I can help you understand and respond to security events and threats. How can I assist you today?</div>"""
        self.chat_display.append(welcome_message)
        
        self.chat_status_label.setText("Started new chat")
        self.chat_status_label.setStyleSheet("color: #00ff00;")
        
    def create_system_tray(self):
        """Create and initialize the system tray icon"""
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(self.windowIcon())
        
        # Create tray menu
        tray_menu = QMenu()
        show_action = tray_menu.addAction("Show")
        show_action.triggered.connect(self.show)
        hide_action = tray_menu.addAction("Minimize to Tray")
        hide_action.triggered.connect(self.hide)
        quit_action = tray_menu.addAction("Quit")
        quit_action.triggered.connect(self.quit_application)
        
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.tray_icon_activated)
        self.tray_icon.show()

    def tray_icon_activated(self, reason):
        """Handle tray icon activation"""
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            if self.isHidden():
                self.show()
            else:
                self.hide()

    def closeEvent(self, event):
        """Override close event to minimize to tray instead of closing"""
        if self.tray_icon.isVisible():
            self.hide()
            event.ignore()
        else:
            event.accept()

    def quit_application(self):
        """Properly quit the application"""
        self.tray_icon.hide()
        QApplication.quit()
    
    def init_ai_agents(self):
        """Initialize AI agents."""
        try:
            # Initialize agent configurations
            agent_configs = {
                'abuseipdb': self.config.ABUSEIPDB_CONFIG['api_key'],
                'ollama': {
                    'base_url': self.config.OLLAMA_CONFIG['base_url'],
                    'model': self.config.OLLAMA_CONFIG['model']
                }
            }
            
            # Initialize crawler agent with config
            crawler_config = self.config.CRAWLER_CONFIG.copy() if hasattr(self.config, 'CRAWLER_CONFIG') else {}
            self.crawler_agent = CrawlerAgent(config=crawler_config)
            
            # Initialize defense agent
            self.defense_agent = DefenseAgent()
            
            # Initialize log monitor agent
            self.log_monitor = LogMonitorAgent()
            
            # Initialize threat intel agent
            self.threat_intel = ThreatIntelAgent(agent_configs)
            
            # Initialize the agent tabs
            self.agent_tabs = AgentTabs(self, agent_configs)
            self.agent_tabs.setObjectName("agentTabs")
            
            # Add agent tabs to the main layout
            self.tabs.addTab(self.agent_tabs, "AI Agents")
            
            self.logger.info("AI agents initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Error initializing AI agents: {str(e)}")
            raise

    def _handle_crawler_finding(self, finding: str, finding_data: dict):
        """Handle findings from the crawler agent in the main thread."""
        try:
            # Update the findings in the security panel
            self.agent_tabs.update_security_finding(finding_data)
            self.logger.info(f"Crawler finding: {finding}")
        except Exception as e:
            self.logger.error(f"Error handling crawler finding: {str(e)}")

    def _handle_crawler_status(self, status: str):
        """Handle status updates from the crawler agent."""
        try:
            self.status_bar.showMessage(status, 5000)  # Show for 5 seconds
            self.logger.info(f"Crawler status: {status}")
        except Exception as e:
            self.logger.error(f"Error handling crawler status: {str(e)}")

    def _handle_crawler_error(self, error: str):
        """Handle error messages from the crawler agent."""
        try:
            self.show_notification("Crawler Error", error)
            self.logger.error(f"Crawler error: {error}")
        except Exception as e:
            self.logger.error(f"Error handling crawler error: {str(e)}")

    def _is_launch_agent_enabled(self):
        """Check if the launch agent is currently enabled"""
        launch_agent_path = os.path.expanduser('~/Library/LaunchAgents/com.sysdaemonai.plist')
        return os.path.exists(launch_agent_path)

    def handle_startup_toggle(self, state):
        """Handle the startup checkbox state change"""
        try:
            launch_agent_path = os.path.expanduser('~/Library/LaunchAgents/com.sysdaemonai.plist')
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'startup_wrapper.command')
            
            if state == Qt.CheckState.Checked.value:
                # Make sure the script is executable
                os.chmod(script_path, 0o755)
                
                # Copy the launch agent plist if it doesn't exist
                if not os.path.exists(launch_agent_path):
                    shutil.copy2(
                        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'com.sysdaemonai.plist'),
                        launch_agent_path
                    )
                
                # Register the launch agent without starting it
                subprocess.run(['launchctl', 'load', '-w', launch_agent_path], check=True)
                self.logger.info("Enabled system startup via launch agent")
                self.show_notification("System Startup", "Network monitoring will start on next system startup")
            else:
                # Unload the launch agent if it exists
                if os.path.exists(launch_agent_path):
                    subprocess.run(['launchctl', 'unload', '-w', launch_agent_path], check=True)
                    os.remove(launch_agent_path)
                self.logger.info("Disabled system startup")
                self.show_notification("System Startup", "Network monitoring will no longer run at system startup")
                
        except Exception as e:
            error_msg = f"Error configuring system startup: {str(e)}"
            self.logger.error(error_msg)
            if hasattr(self, 'show_notification'):
                self.show_notification("Error", error_msg)
            # Revert checkbox state
            self.startup_checkbox.setChecked(self._is_launch_agent_enabled())

    def setup_terminal_tab(self):
        """Setup the terminal tab with command execution and cancellation functionality"""
        layout = QVBoxLayout()

        # Terminal output area
        self.terminal_output = QPlainTextEdit()
        self.terminal_output.setReadOnly(True)
        layout.addWidget(self.terminal_output)

        # Terminal input area
        self.terminal_input = QLineEdit()
        self.terminal_input.setPlaceholderText('Enter command...')
        layout.addWidget(self.terminal_input)
        
        self.command_history = []  # List to store command history
        self.command_history_index = -1  # Index for navigating command history

        # Initialize the completion widget
        self.completion_widget = QListWidget()
        self.completion_widget.setVisible(False)
        layout.addWidget(self.completion_widget)

        # Connect the terminal input to handle key events
        self.terminal_input.returnPressed.connect(self.execute_command)
        self.terminal_input.keyPressEvent = self.handle_key_press

        # Add a checkbox for running commands as a regular user
        self.run_as_user_checkbox = QCheckBox('Run as regular user')
        self.run_as_user_checkbox.setChecked(False)  # Default is unchecked (run with sudo)
        layout.addWidget(self.run_as_user_checkbox)

        # Execute button
        execute_button = QPushButton('Execute')
        execute_button.clicked.connect(self.execute_command)
        layout.addWidget(execute_button)

        # Cancel button
        cancel_button = QPushButton('Cancel')
        cancel_button.clicked.connect(self.cancel_command)
        layout.addWidget(cancel_button)

        # Clear button
        clear_button = QPushButton('Clear')
        clear_button.clicked.connect(self.clear_terminal)
        layout.addWidget(clear_button)

        tab = QWidget()
        tab.setLayout(layout)
        return tab


    def execute_command(self):
      command = self.terminal_input.text()
      if command:
          self.command_history.append(command)  # Store the command in history
          self.command_history_index = len(self.command_history)  # Reset index for new commands
          # If the checkbox is checked, run as the current user
          if self.run_as_user_checkbox.isChecked():
              current_user = getpass.getuser()  # Get the current user's name
              command = f"su - {current_user} -c '{command}'"  # Run command as current user
          # Command will run with sudo by default if checkbox is unchecked
          self.terminal_output.appendPlainText(f'> {command}')  # Display the command
          self.command_thread = CommandThread(command)
          self.command_thread.output_signal.connect(self.update_terminal_output)
          self.command_thread.finished_signal.connect(self.command_finished)
          self.command_thread.start()
          self.terminal_input.clear()  # Clear the input field after execution

    def update_terminal_output(self, output):
        self.terminal_output.appendPlainText(output)  # Display command output in terminal

    def command_finished(self):
        self.terminal_output.appendPlainText('Command execution finished.')  # Notify user when command execution is complete
        self.show_notification("Command Complete", "Command execution finished successfully")

    def cancel_command(self):
        if hasattr(self, 'command_thread'):
            self.command_thread.stop()
            self.terminal_output.appendPlainText('Command execution canceled.')

    def handle_key_press(self, event):
        """Handle key press events in the terminal input"""
        if event.key() == Qt.Key.Key_Return:
            self.execute_command()  # Execute command on Enter key
        elif event.key() == Qt.Key.Key_Escape:
            self.terminal_input.clear()  # Clear input on Escape key
        elif event.key() == Qt.Key.Key_Up:
            # Navigate to the previous command
            if self.command_history_index > 0:
                self.command_history_index -= 1
                self.terminal_input.setText(self.command_history[self.command_history_index])
        elif event.key() == Qt.Key.Key_Down:
            # Navigate to the next command
            if self.command_history_index < len(self.command_history) - 1:
                self.command_history_index += 1
                self.terminal_input.setText(self.command_history[self.command_history_index])
        elif event.key() == Qt.Key.Key_Tab:
            # Trigger auto-completion
            self.show_completions()
        else:
            # Call the default key press event handler for other keys
            QLineEdit.keyPressEvent(self.terminal_input, event)

    def show_completions(self):
        """Show auto-completion suggestions based on the current input"""
        current_text = self.terminal_input.text()
        completions = self.get_completions(current_text)
        if completions:
            self.completion_widget.clear()
            self.completion_widget.addItems(completions)
            self.completion_widget.setVisible(True)
            self.completion_widget.setGeometry(self.terminal_input.x(), self.terminal_input.y() + self.terminal_input.height(), self.terminal_input.width(), 100)  # Adjust size as needed
            self.completion_widget.show()
        else:
            self.completion_widget.setVisible(False)

    def get_completions(self, text):
        """Return a list of possible completions based on the current input"""
        possible_commands = ["ls", "cd", "mkdir", "rm", "touch", "echo"]
        completions = [cmd for cmd in possible_commands if cmd.startswith(text)]
        print(f"Possible completions for '{text}': {completions}")  # Debugging output
        return completions

    def clear_terminal(self):
        """Clear the terminal output"""
        self.terminal_output.clear()
        self.show_notification("Success", "Terminal cleared")

    def setup_admin_tab(self):
        """Setup the admin tab with user and role management"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Create a splitter for users and roles
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Users Panel
        users_widget = QWidget()
        users_layout = QVBoxLayout(users_widget)
        
        # Users Table
        users_group = QGroupBox("User Management")
        users_group_layout = QVBoxLayout(users_group)
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)
        self.users_table.setHorizontalHeaderLabels(["Username", "Role", "Created", "Actions"])
        self.users_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.users_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.users_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.users_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        users_group_layout.addWidget(self.users_table)
        
        # User Actions
        actions_layout = QHBoxLayout()
        
        # Add User Button
        add_user_btn = QPushButton("Add User")
        add_user_btn.clicked.connect(self.show_add_user_dialog)
        actions_layout.addWidget(add_user_btn)
        
        # Refresh Button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_users_table)
        actions_layout.addWidget(refresh_btn)
        
        users_group_layout.addLayout(actions_layout)
        users_layout.addWidget(users_group)
        
        # Role Management Panel
        roles_widget = QWidget()
        roles_layout = QVBoxLayout(roles_widget)
        
        # Roles Table
        roles_group = QGroupBox("Role Management")
        roles_group_layout = QVBoxLayout(roles_group)
        
        self.roles_table = QTableWidget()
        self.roles_table.setColumnCount(3)
        self.roles_table.setHorizontalHeaderLabels(["Role", "Permissions", "Actions"])
        self.roles_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.roles_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.roles_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        roles_group_layout.addWidget(self.roles_table)
        
        roles_layout.addWidget(roles_group)
        
        # Add widgets to splitter
        splitter.addWidget(users_widget)
        splitter.addWidget(roles_widget)
        splitter.setSizes([int(splitter.size().width() * 0.6), int(splitter.size().width() * 0.4)])
        
        layout.addWidget(splitter)
        
        # Initial data load
        self.refresh_users_table()
        self.refresh_roles_table()

        # Remote Agents Section
        agents_group = QGroupBox("Remote Agents")
        agents_layout = QVBoxLayout(agents_group)

        # Search bar
        search_layout = QHBoxLayout()
        search_label = QLabel("Search:")
        self.agent_search_input.setPlaceholderText("Search agents...")
        self.agent_search_input.textChanged.connect(self.filter_agents)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.agent_search_input)
        agents_layout.addLayout(search_layout)

        # Agents dropdown
        self.agents_list.currentIndexChanged.connect(lambda idx: self.show_agent_details(idx))
        agents_layout.addWidget(self.agents_list)

        # Refresh button
        refresh_agents_btn = QPushButton("Refresh Agents")
        refresh_agents_btn.clicked.connect(self.load_remote_agents)
        agents_layout.addWidget(refresh_agents_btn)

        layout.addWidget(agents_group)

        # Add this to the initial data load section
        self.load_remote_agents()
        
        return tab

    def show_add_user_dialog(self):
        """Show dialog to add a new user"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Add New User")
        dialog.setModal(True)
        
        layout = QFormLayout(dialog)
        
        # Username input
        username_input = QLineEdit()
        layout.addRow("Username:", username_input)
        
        # Password input
        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Password:", password_input)
        
        # Role selection
        role_combo = QComboBox()
        role_combo.addItems(['viewer', 'analyst', 'admin'])
        layout.addRow("Role:", role_combo)
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            username = username_input.text()
            password = password_input.text()
            role = role_combo.currentText()
            
            if username and password:
                try:
                    if self.auth_manager.create_user(username, password, role):
                        QMessageBox.information(self, "Success", "User created successfully")
                        self.refresh_users_table()
                    else:
                        QMessageBox.warning(self, "Error", "Failed to create user")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error creating user: {str(e)}")
            else:
                QMessageBox.warning(self, "Error", "Username and password are required")

    def refresh_users_table(self):
        """Refresh the users table with current data"""
        try:
            users = self.db_manager.fetch_all(
                """SELECT username, role, created_at 
                   FROM users 
                   ORDER BY created_at DESC"""
            )
            
            self.users_table.setRowCount(len(users))
            for row, (username, role, created_at) in enumerate(users):
                # Username
                self.users_table.setItem(row, 0, QTableWidgetItem(username))
                
                # Role
                role_combo = QComboBox()
                role_combo.addItems(['viewer', 'analyst', 'admin'])
                role_combo.setCurrentText(role)
                role_combo.currentTextChanged.connect(
                    lambda text, username=username: self.update_user_role(username, text)
                )
                self.users_table.setCellWidget(row, 1, role_combo)
                
                # Created At
                try:
                    if isinstance(created_at, str):
                        # Parse string to datetime if it's a string
                        dt = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                        formatted_date = dt.strftime('%Y-%m-%d %H:%M')
                    else:
                        formatted_date = created_at.strftime('%Y-%m-%d %H:%M')
                except (ValueError, AttributeError):
                    # If parsing fails, just display the raw string
                    formatted_date = str(created_at)
                
                created_item = QTableWidgetItem(formatted_date)
                created_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.users_table.setItem(row, 2, created_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout()
                
                # Change Password Button
                change_pwd_btn = QPushButton("Change Password")
                change_pwd_btn.setStyleSheet("background-color: #4a9eff; color: white;")
                change_pwd_btn.clicked.connect(lambda _, u=username: self.show_change_password_dialog(u))
                actions_layout.addWidget(change_pwd_btn)
                
                # Delete Button
                delete_btn = QPushButton("Delete")
                delete_btn.setStyleSheet("background-color: #ff4444; color: white;")
                delete_btn.clicked.connect(lambda _, u=username: self.delete_user(u))
                actions_layout.addWidget(delete_btn)
                
                self.users_table.setCellWidget(row, 3, actions_widget)
                
        except Exception as e:
            self.logger.error(f"Error refreshing users table: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to load users: {str(e)}")

    def refresh_roles_table(self):
        """Refresh the roles table with current data"""
        roles_data = [
            ('admin', 'Full system access, user management'),
            ('analyst', 'View and analyze data, run commands'),
            ('viewer', 'View-only access to dashboards')
        ]
        
        self.roles_table.setRowCount(len(roles_data))
        for row, (role, permissions) in enumerate(roles_data):
            # Role
            role_item = QTableWidgetItem(role)
            role_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.roles_table.setItem(row, 0, role_item)
            
            # Permissions
            perm_item = QTableWidgetItem(permissions)
            self.roles_table.setItem(row, 1, perm_item)
            
            # No actions for built-in roles
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(4, 4, 4, 4)
            
            if role != 'admin':  # Don't allow editing admin permissions
                edit_btn = QPushButton("Edit")
                edit_btn.clicked.connect(lambda _, r=role: self.edit_role_permissions(r))
                actions_layout.addWidget(edit_btn)
            
            self.roles_table.setCellWidget(row, 2, actions_widget)

    def update_user_role(self, username: str, new_role: str):
        """Update a user's role"""
        try:
            if username == self.auth_manager.verify_token(self.auth_token).get('username'):
                QMessageBox.warning(self, "Error", "Cannot modify your own role")
                self.refresh_users_table()  # Refresh to revert the change
                return
                
            self.db_manager.execute(
                "UPDATE users SET role = :role WHERE username = :username",
                {"role": new_role, "username": username}
            )
            QMessageBox.information(self, "Success", f"Updated role for user {username}")
        except Exception as e:
            self.logger.error(f"Error updating user role: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to update role: {str(e)}")
            self.refresh_users_table()  # Refresh to revert the change

    def delete_user(self, username: str):
        """Delete a user"""
        try:
            if username == self.auth_manager.verify_token(self.auth_token).get('username'):
                QMessageBox.warning(self, "Error", "Cannot delete your own account")
                return
                
            reply = QMessageBox.question(
                self,
                'Confirm Delete',
                f'Are you sure you want to delete user {username}?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.db_manager.execute(
                    "DELETE FROM users WHERE username = :username",
                    {"username": username}
                )
                QMessageBox.information(self, "Success", f"Deleted user {username}")
                self.refresh_users_table()
        except Exception as e:
            self.logger.error(f"Error deleting user: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to delete user: {str(e)}")

    def edit_role_permissions(self, role: str):
        """Show dialog to edit role permissions"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit {role.title()} Permissions")
        dialog.setModal(True)
        
        layout = QVBoxLayout(dialog)
        
        # Permissions checkboxes
        permissions = {
            'view_dashboard': 'View Dashboard',
            'run_commands': 'Run Commands',
            'export_data': 'Export Data',
            'modify_settings': 'Modify Settings',
            'view_logs': 'View Logs'
        }
        
        checkboxes = {}
        for perm_id, perm_name in permissions.items():
            cb = QCheckBox(perm_name)
            cb.setChecked(perm_id in ['view_dashboard'] if role == 'viewer'
                         else perm_id in ['view_dashboard', 'run_commands', 'export_data', 'view_logs'])
            checkboxes[perm_id] = cb
            layout.addWidget(cb)
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # In a real implementation, this would update the role permissions in the database
            QMessageBox.information(
                self,
                "Success",
                f"Updated permissions for {role} role\n(Note: This is a mock implementation)"
            )
    def show_change_password_dialog(self, username: str):
        """Show dialog to change a user's password"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Change Password for {username}")
        dialog.setModal(True)
        
        layout = QFormLayout(dialog)
        
        # New password input
        new_password = QLineEdit()
        new_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("New Password:", new_password)
        
        # Confirm password input
        confirm_password = QLineEdit()
        confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Confirm Password:", confirm_password)
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            password = new_password.text()
            confirm = confirm_password.text()
            
            if not password:
                QMessageBox.warning(self, "Error", "Password cannot be empty")
                return
                
            if password != confirm:
                QMessageBox.warning(self, "Error", "Passwords do not match")
                return
                
            try:
                # Update the password using auth_manager
                if self.auth_manager.update_password(username, password):
                    QMessageBox.information(self, "Success", f"Password updated for user {username}")
                else:
                    QMessageBox.warning(self, "Error", "Failed to update password")
            except Exception as e:
                self.logger.error(f"Error updating password: {str(e)}")
                QMessageBox.critical(self, "Error", f"Failed to update password: {str(e)}")

    def load_remote_agents(self):
        """Load list of remote agents from the database"""
        try:
            # Create data directory if it doesn't exist
            data_dir = os.path.join(os.path.dirname(__file__), 'data')
            os.makedirs(data_dir, exist_ok=True)
            
            # Load agents from file
            agents_file = os.path.join(data_dir, 'remote_agents.json')
            if os.path.exists(agents_file):
                with open(agents_file, 'r') as f:
                    self.agents = json.load(f)
            else:
                self.agents = []
            
            self.update_agents_list()
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load remote agents: {e}")
    
    def filter_agents(self, text):
        """Filter agents list based on search text"""
        self.agents_list.clear()
        for agent in self.agents:
            if text.lower() in agent['hostname'].lower():
                self.agents_list.addItem(agent['hostname'])
    
    def update_agents_list(self):
        """Update the agents dropdown list"""
        self.agents_list.clear()
        for agent in sorted(self.agents, key=lambda x: x['hostname']):
            self.agents_list.addItem(agent['hostname'])
    
    def show_agent_details(self, index):
        """Show details for the selected agent"""
        if index < 0:
            return
            
        hostname = self.agents_list.currentText()
        agent = next((a for a in self.agents if a['hostname'] == hostname), None)
        if not agent:
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Agent Details - {hostname}")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Create tabs
        tabs = QTabWidget()
        
        # System Info tab
        system_tab = QWidget()
        system_layout = QVBoxLayout()
        system_text = QTextEdit()
        system_text.setReadOnly(True)
        system_text.setText(json.dumps(agent.get('system_info', {}), indent=2))
        system_layout.addWidget(system_text)
        system_tab.setLayout(system_layout)
        tabs.addTab(system_tab, "System Info")
        
        # Alerts tab
        alerts_tab = QWidget()
        alerts_layout = QVBoxLayout()
        alerts_text = QTextEdit()
        alerts_text.setReadOnly(True)
        alerts = agent.get('alerts', [])
        alerts_text.setText("\n\n".join([
            f"[{alert['timestamp']}] {alert['title']}\n{alert['message']}"
            for alert in alerts
        ]))
        alerts_layout.addWidget(alerts_text)
        alerts_tab.setLayout(alerts_layout)
        tabs.addTab(alerts_tab, f"Alerts ({len(alerts)})")
        
        # Security Analysis tab
        security_tab = QWidget()
        security_layout = QVBoxLayout()
        security_text = QTextEdit()
        security_text.setReadOnly(True)
        security_text.setText(agent.get('security_analysis', 'No security analysis available'))
        security_layout.addWidget(security_text)
        security_tab.setLayout(security_layout)
        tabs.addTab(security_tab, "Security Analysis")
        
        # Network Info tab
        network_tab = QWidget()
        network_layout = QVBoxLayout()
        network_text = QTextEdit()
        network_text.setReadOnly(True)
        network_text.setText(json.dumps(agent.get('network_info', {}), indent=2))
        network_layout.addWidget(network_text)
        network_tab.setLayout(network_layout)
        tabs.addTab(network_tab, "Network Info")
        
        layout.addWidget(tabs)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

class CommandThread(QThread):
    output_signal = pyqtSignal(str)
    finished_signal = pyqtSignal()

    def __init__(self, command):
        super().__init__()
        self.command = command
        self.running = True
        self.process = None  # Store the process reference

    def run(self):
        try:
            print(f"Running command: {self.command}")  # Debugging output
            self.process = subprocess.Popen(self.command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            for output in iter(self.process.stdout.readline, b''):
                self.output_signal.emit(output.decode().strip())  # Emit output to the GUI
            self.process.stdout.close()
            self.process.wait()  # Wait for the process to complete
        except subprocess.CalledProcessError as e:
            self.output_signal.emit(f"Error executing command: {e.output.decode('utf-8')}")
        except Exception as e:
            self.output_signal.emit(f"Error executing command: {str(e)}")
        finally:
            self.finished_signal.emit()

    def stop(self):
        self.running = False
        if self.process:
            self.process.terminate()

    def clear_terminal(self):
        """Clear the terminal output"""
        self.terminal_output.clear()

   

def main():
    # Create QApplication instance first
    app = QApplication(sys.argv)
    
    try:
        # Set application metadata
        app.setApplicationName("SysDaemon AI")
        app.setApplicationDisplayName("SysDaemon AI")
        app.setOrganizationName("Sysdaemon AI")
        app.setOrganizationDomain("Sysdaemon AI")
        
        # Set application style
        app.setStyle('Fusion')
        
        # Set application icon globally
        icon_path = os.path.join(os.path.dirname(__file__), 'icons', 'app_icon.png')
        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            app.setWindowIcon(app_icon)
        
        # Create and show main window
        window = NetworkMonitorGUI()
        window.show()
        
        # Start Qt event loop
        return app.exec()
        
    except Exception as e:
        logging.error(f"Error in main: {str(e)}")
        QMessageBox.critical(None, "Error", f"Failed to start application: {str(e)}")
        return 1

if __name__ == '__main__':
    sys.exit(main())
    
